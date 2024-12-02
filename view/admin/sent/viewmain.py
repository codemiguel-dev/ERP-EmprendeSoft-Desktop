import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllersent import SentController
from view.admin.sent.viewadd import Viewadd
from view.admin.sent.viewupdate import Viewupdate


class Viewmainsent(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmainsent, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/admin/mainsent{self.theme}.ui", self)

        icon_excel(self)
        icon_exit_program(self)
        icon_configurate_manager(self)
        icon_configurate_top(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show_sent)
        self.btn_update.clicked.connect(self.update)
        self.btn_delete.clicked.connect(self.delete)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = SentController(self)

        self.show_sent()

    def close_program(self):
        QApplication.quit()

    def reload_ui(self):
        self.close()  # Cierra la ventana actual
        self.__init__()  # Re-inicia la vista con el nuevo tema
        self.show()  # Vuelve a mostrar la ventana

    def add(self):
        self.inventory_add = Viewadd()
        self.inventory_add.show()

    def show_sent(self):
        # Obtener datos de usuarios desde el controlador
        Sent_controller = self.controller.get()

        # print(Client_controller)  # Añade esta línea para depurar

        # Limpiar la tabla
        self.table_sent.setRowCount(0)

        self.table_sent.verticalHeader().setDefaultSectionSize(200)
        self.table_sent.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar inventario en la tabla
        for i, sent in enumerate(Sent_controller):
            self.table_sent.insertRow(i)

            # Verifica si inventory es una tupla o una lista de un solo valor
            if isinstance(sent, (tuple, list)):

                # Aquí asumimos que es una tupla o lista con múltiples elementos
                self.table_sent.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(sent[0]))
                )  # ID

                self.table_sent.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(sent[1])
                )  # METHOD
                self.table_sent.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(sent[2])
                )  # DESCRIPTION
                self.table_sent.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(str(sent[3]))
                )  # PRICE
                self.table_sent.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(sent[4])
                )  # STATUS
                self.table_sent.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(sent[5])
                )  # COUNTRY
                self.table_sent.setItem(
                    i, 6, QtWidgets.QTableWidgetItem(sent[6])
                )  # COMMUNE
                self.table_sent.setItem(
                    i, 7, QtWidgets.QTableWidgetItem(sent[7])
                )  # REGION
                self.table_sent.setItem(
                    i, 8, QtWidgets.QTableWidgetItem(sent[8])
                )  # DESCRIPTION ADDRESS
            else:
                # Si inventory es un valor único, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único:sentnt"
                )

    def update(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_sent.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_sent.item(selected_row, 0)
            method_item = self.table_sent.item(selected_row, 1)  # method
            description_item = self.table_sent.item(selected_row, 2)  # description
            price_sent_item = self.table_sent.item(selected_row, 3)  # price sent
            status_sent_item = self.table_sent.item(selected_row, 4)  # status sent
            country_item = self.table_sent.item(selected_row, 5)  # country
            region_item = self.table_sent.item(selected_row, 6)  # region
            commune_item = self.table_sent.item(selected_row, 7)  # commune
            address_description_item = self.table_sent.item(selected_row, 8)  # address

            if (
                uid_item
                and method_item
                and description_item
                and price_sent_item
                and status_sent_item
                and country_item
                and region_item
                and commune_item
                and address_description_item
            ):
                uid = uid_item.text()
                method = method_item.text()
                description = description_item.text()
                price_sent = price_sent_item.text()
                status = status_sent_item.text()
                country = country_item.text()
                region = region_item.text()
                commune = commune_item.text()
                address_description = address_description_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(
                    uid,
                    method,
                    description,
                    price_sent,
                    status,
                    country,
                    region,
                    commune,
                    address_description,
                )
                self.update_form.show()
                self.show()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def delete(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_sent.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_sent.item(selected_row, 0)

            if uid_item:
                uid = uid_item.text()

                # Cuadro de diálogo de confirmación
                reply = QtWidgets.QMessageBox.question(
                    self,
                    "Confirmar eliminación",
                    f"¿Estás seguro de que deseas eliminar el usuario con ID {uid}?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No,
                )

                # Si el usuario confirma, se procede a eliminar
                if reply == QtWidgets.QMessageBox.Yes:
                    self.controller.delete(uid)
                    self.show_sent()
                else:
                    # Cancelar la eliminación
                    QtWidgets.QMessageBox.information(
                        self, "Cancelado", "La eliminación ha sido cancelada."
                    )

        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        sent_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            sent_controller,
            columns=[
                "ID",
                "Metodo",
                "Descripción",
                "Precio de envío",
                "Estado del envío",
                "Pais",
                "Region",
                "Comuna",
                "Descripción de dirección",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Envío exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_sent.rowCount()):
            match = False
            for col in range(self.table_sent.columnCount()):
                item = self.table_sent.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_sent.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
