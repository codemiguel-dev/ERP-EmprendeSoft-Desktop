import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllerprovider import ProviderController
from view.admin.provider.viewadd import Viewadd
from view.admin.provider.viewupdate import Viewupdate


class Viewmainprovider(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmainprovider, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/standar/mainprovider{self.theme}.ui", self)

        # icon_configurate_manager(self)
        icon_configurate_top(self)
        icon_exit_program(self)
        icon_excel(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show_provider)
        self.btn_update.clicked.connect(self.update)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = ProviderController(self)

        self.show_provider()

    def close_program(self):
        QApplication.quit()

    def add(self):
        self.provider_add = Viewadd()
        self.provider_add.show()

    def show_provider(self):
        # Obtener datos de usuarios desde el controlador
        Provider_controller = self.controller.get()

        # Limpiar la tabla
        self.table_provider.setRowCount(0)

        self.table_provider.verticalHeader().setDefaultSectionSize(200)
        self.table_provider.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar inventario en la tabla
        for i, provider in enumerate(Provider_controller):
            self.table_provider.insertRow(i)

            # Verifica si inventory es una tupla o una lista de un solo valor
            if isinstance(provider, (tuple, list)):
                # ID
                self.table_provider.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(provider[0]))
                )

                # Imagen (asumiendo que está en la columna 1, posición de "name")
                if provider[1] is not None:  # Asegúrate de que el BLOB no sea nulo
                    pixmap = QPixmap()
                    pixmap.loadFromData(provider[1])  # Convierte el BLOB en una imagen
                    image_label = QtWidgets.QLabel()  # Crea un QLabel para la imagen
                    image_label.setPixmap(
                        pixmap.scaled(200, 200, QtCore.Qt.KeepAspectRatio)
                    )  # Ajusta el tamaño de la imagen
                    self.table_provider.setCellWidget(
                        i, 1, image_label
                    )  # Muestra la imagen en la tabla
                else:
                    self.table_provider.setItem(
                        i, 1, QtWidgets.QTableWidgetItem("No image")
                    )  # Muestra texto si no hay imagen

                # Otros datos
                self.table_provider.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(provider[2])
                )  # Rut
                self.table_provider.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(provider[3])
                )  # Name
                self.table_provider.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(provider[4])
                )  # Email
                self.table_provider.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(provider[5])
                )  # Contact
                self.table_provider.setItem(
                    i, 6, QtWidgets.QTableWidgetItem(provider[6])
                )  # Address
                self.table_provider.setItem(
                    i, 7, QtWidgets.QTableWidgetItem(provider[7])
                )  # Type provider
            else:
                # Si inventory es un valor único, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {provider}"
                )

    def update(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_provider.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_provider.item(selected_row, 0)
            image_item = self.table_provider.item(selected_row, 1)
            rut_item = self.table_provider.item(selected_row, 2)
            name_item = self.table_provider.item(selected_row, 3)
            email_item = self.table_provider.item(selected_row, 4)
            fono_item = self.table_provider.item(selected_row, 5)
            address_item = self.table_provider.item(selected_row, 6)
            type_item = self.table_provider.item(selected_row, 7)

            if (
                uid_item
                and rut_item
                and name_item
                and email_item
                and fono_item
                and address_item
                and type_item
            ):
                uid = uid_item.text()
                rut = rut_item.text()
                name = name_item.text()
                email = email_item.text()
                fono = fono_item.text()
                address = address_item.text()
                type_provider = type_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(
                    uid, rut, name, email, fono, address, type_provider
                )
                self.update_form.show()
                self.show_provider()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        provider_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            provider_controller,
            columns=[
                "ID",
                "Imagen",
                "Rut",
                "Nombre",
                "Correo",
                "Tel. Contacto",
                "Dirección",
                "Tipo de Proveedor",
            ],
        )

        # Validar si hay imágenes nulas
        if df["Imagen"].isnull().any():
            QtWidgets.QMessageBox.warning(
                self, "Advertencia", "No hay imagen para algunos usuarios."
            )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_provider.rowCount()):
            match = False
            for col in range(self.table_provider.columnCount()):
                item = self.table_provider.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_provider.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
