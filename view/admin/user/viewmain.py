import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QFileDialog, QMessageBox, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controlleruser import UserController
from view.admin.user.viewadd import Viewadd
from view.admin.user.viewupdate import Viewupdate


class Viewmainuser(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmainuser, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/admin/mainuser{self.theme}.ui", self)

        icon_exit_program(self)
        icon_excel(self)
        icon_configurate_manager(self)
        icon_configurate_top(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show_user)
        self.btn_update.clicked.connect(self.update)
        self.btn_delete.clicked.connect(self.delete)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = UserController(self)

        self.show_user()

    def close_program(self):
        QApplication.quit()

    def add(self):
        self.inventory_add = Viewadd()
        self.inventory_add.show()

    def show_user(self):
        # Obtener datos de usuarios desde el controlador
        User_controller = self.controller.get_user()

        # Limpiar la tabla
        self.table_user.setRowCount(0)

        self.table_user.verticalHeader().setDefaultSectionSize(200)
        self.table_user.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar inventario en la tabla
        for i, user in enumerate(User_controller):
            self.table_user.insertRow(i)

            # Verifica si inventory es una tupla o una lista de un solo valor
            if isinstance(user, (tuple, list)):
                # Aquí asumimos que es una tupla o lista con múltiples elementos
                self.table_user.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(user[0]))
                )  # ID

                # Imagen (asumiendo que está en la columna 1, posición de "name")
                if user[1] is not None:  # Asegúrate de que el BLOB no sea nulo
                    pixmap = QPixmap()
                    pixmap.loadFromData(user[1])  # Convierte el BLOB en una imagen
                    image_label = QtWidgets.QLabel()  # Crea un QLabel para la imagen
                    image_label.setPixmap(
                        pixmap.scaled(200, 200, QtCore.Qt.KeepAspectRatio)
                    )  # Ajusta el tamaño de la imagen
                    self.table_user.setCellWidget(
                        i, 1, image_label
                    )  # Muestra la imagen en la tabla
                else:
                    self.table_user.setItem(
                        i, 1, QtWidgets.QTableWidgetItem("No image")
                    )  # Muestra texto si no hay imagen
                self.table_user.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(user[2])
                )  # Name
                self.table_user.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(str(user[3]))
                )  # Password
                self.table_user.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(user[4])
                )  # Email
                self.table_user.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(user[5])
                )  # Contact
                self.table_user.setItem(
                    i, 6, QtWidgets.QTableWidgetItem(user[6])
                )  # Address
                self.table_user.setItem(
                    i, 7, QtWidgets.QTableWidgetItem(user[7])
                )  # Role
            else:
                # Si inventory es un valor único, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {user}"
                )

    def update(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_user.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_user.item(selected_row, 0)
            image_item = self.table_user.item(selected_row, 1)
            name_item = self.table_user.item(selected_row, 2)
            password_item = self.table_user.item(selected_row, 3)
            email_item = self.table_user.item(selected_row, 4)
            fono_item = self.table_user.item(selected_row, 5)
            address_item = self.table_user.item(selected_row, 6)
            role_item = self.table_user.item(selected_row, 7)

            if (
                uid_item
                and name_item
                and email_item
                and fono_item
                and address_item
                and password_item
            ):
                uid = uid_item.text()
                name = name_item.text()
                email = email_item.text()
                fono = fono_item.text()
                address = address_item.text()
                password = password_item.text()
                role = role_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(
                    uid, name, email, fono, address, password, role
                )
                self.update_form.show()
                self.show_user()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def delete(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_user.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_user.item(selected_row, 0)

            if uid_item:
                uid = uid_item.text()

                # Cuadro de diálogo de confirmación
                reply = QtWidgets.QMessageBox.question(
                    self,
                    "Confirmar eliminación",
                    f"¿Estás seguro de que deseas eliminar el usuario con ID {uid}?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No,
                )

                # Si el usuario confirma, se procede a eliminar
                if reply == QtWidgets.QMessageBox.Yes:
                    self.controller.delete_user(uid)
                    self.show_user()
                else:
                    # Cancelar la eliminación
                    QtWidgets.QMessageBox.information(
                        self, "Cancelado", "La eliminación ha sido cancelada."
                    )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para eliminar."
            )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        user_controller = self.controller.get_user()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            user_controller,
            columns=[
                "ID",
                "Imagen",
                "Nombre",
                "Contraseña HASH",
                "Correo",
                "Teléfono",
                "Dirección",
                "Rol del usuario",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            # Abrir el archivo en Excel
            os.startfile(file_path)

            # Mostrar mensaje de éxito
            QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            # Mostrar mensaje de cancelación
            QMessageBox.warning(self, "Cancelado", "La exportación fue cancelada.")

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_user.rowCount()):
            match = False
            for col in range(self.table_user.columnCount()):
                item = self.table_user.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_user.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
