import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllercalendar import CalendarController
from view.admin.calendar.viewadd import Viewadd
from view.admin.calendar.viewupdate import Viewupdate


class Viewmaincalendar(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmaincalendar, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/standar/maincalendar{self.theme}.ui", self)

        icon_configurate_top(self)
        icon_excel(self)
        icon_exit_program(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_get.clicked.connect(self.show)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = CalendarController(self)

        self.show()

    def close_program(self):
        QApplication.quit()

    def show(self):
        # Obtener datos de usuarios desde el controlador
        Calendar_controller = self.controller.get()

        # print(Client_controller)  # Añade esta línea para depurar

        # Limpiar la tabla
        self.table_calendar.setRowCount(0)

        self.table_calendar.verticalHeader().setDefaultSectionSize(200)
        self.table_calendar.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar inventario en la tabla
        for i, calendar in enumerate(Calendar_controller):
            self.table_calendar.insertRow(i)

            # Aquí asumimos que es una tupla o lista con múltiples elementos
            self.table_calendar.setItem(
                i, 0, QtWidgets.QTableWidgetItem(calendar[0])
            )  # ID

            self.table_calendar.setItem(
                i, 1, QtWidgets.QTableWidgetItem(calendar[1])
            )  # Name
            self.table_calendar.setItem(
                i, 2, QtWidgets.QTableWidgetItem(str(calendar[2]))
            )  # Last name
            self.table_calendar.setItem(
                i, 3, QtWidgets.QTableWidgetItem(calendar[3])
            )  # Email
            self.table_calendar.setItem(
                i, 4, QtWidgets.QTableWidgetItem(calendar[4])
            )  # Phone
            self.table_calendar.setItem(
                i, 5, QtWidgets.QTableWidgetItem(calendar[5])
            )  # Phone

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        horary_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            horary_controller,
            columns=[
                "Nombre del trabajador",
                "Puesto del trabajo",
                "id del horario",
                "Hora de inicio",
                "Hora de termino",
                "Días de trabajo",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_calendar.rowCount()):
            match = False
            for col in range(self.table_calendar.columnCount()):
                item = self.table_calendar.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_calendar.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
