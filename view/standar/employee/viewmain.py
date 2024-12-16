import os

import pandas as pd
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controlleremployee import EmployeeController
from view.admin.employee.viewadd import Viewadd
from view.admin.employee.viewupdate import Viewupdate
from view.admin.maps.viewmaps import MapaApp


class Viewmainemployee(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmainemployee, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/standar/mainemployee{self.theme}.ui", self)

        icon_configurate_top(self)
        icon_exit_program(self)
        icon_excel(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_get.clicked.connect(self.show)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_maps.clicked.connect(self.maps)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = EmployeeController(self)

        self.show()

    def close_program(self):
        QApplication.quit()

    def maps(self):
        """Obtiene las coordenadas de latitud y longitud para una dirección dada y muestra el mapa."""
        geolocator = Nominatim(user_agent="mi_aplicacion_geocodificador")
        address = "Avenida Ignacio Carrera Pinto 397, Illapel, Coquimbo, Chile"

        try:
            # Geocodificar la dirección
            location = geolocator.geocode(address)

            # Verificar si se obtuvo una ubicación
            if location:
                print(f"Coordenadas de '{address}':")
                print(f"Latitud: {location.latitude}, Longitud: {location.longitude}")
                self.lat = location.latitude
                self.lon = location.longitude

                # Inicializar y mostrar el mapa con las coordenadas
                self.map = MapaApp(self.lat, self.lon)
                self.map.show()
            else:
                print(f"No se encontró la dirección: {address}")

        except Exception as e:
            print(f"Ocurrió un error durante la geocodificación: {e}")

    def show(self):
        # Obtener datos desde el controlador (usando elemployeeJOIN)
        Employee_controller = self.controller.get()

        # Limpiar la tabla antes de insertar nuevos datos
        self.table_employee.setRowCount(0)

        # Configuración del tamaño de las filas y columnas
        self.table_employee.verticalHeader().setDefaultSectionSize(200)
        self.table_employee.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar los datos obtenidos en la tabla
        for i, employee in enumerate(Employee_controller):
            self.table_employee.insertRow(i)

            # Verifica si `task` es una tupla o lista con múltiples elementos
            if isinstance(employee, (tuple, list)):
                # Asumiendo que la consulta INNER JOIN devuelve al menos estos campos:
                # (user_id, user_name, task_id, task_description, task_status)
                self.table_employee.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(employee[0]))
                )  # user_id
                self.table_employee.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(employee[1])
                )  # country
                self.table_employee.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(str(employee[2]))
                )  # region
                self.table_employee.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(employee[3])
                )  # region
            else:
                # Si `task` no es una tupla o lista, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {employee}"
                )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        employee_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            employee_controller,
            columns=[
                "ID del usuario",
                "Nombre",
                "ID del puesto",
                "Puesto del trabajo",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_employee.rowCount()):
            match = False
            for col in range(self.table_employee.columnCount()):
                item = self.table_employee.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_employee.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
