import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_exit_session,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllerinventory import InventoryController
from view.admin.inventory.viewadd import Viewadd
from view.admin.inventory.viewupdate import Viewupdate


class Viewmaininventory(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmaininventory, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/standar/maininventory{self.theme}.ui", self)

        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))

        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        icon_configurate_top(self)
        icon_excel(self)
        icon_exit_program(self)
        icon_configurate_exit_session(self)

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show_inventory)
        self.btn_update.clicked.connect(self.update_inventory)
        self.btn_excel.clicked.connect(self.export_excel_inventory)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)
        self.controller = InventoryController(self)
        self.show_inventory()

    def close_program(self):
        QApplication.quit()

    def add(self):
        self.inventory_add = Viewadd()
        self.inventory_add.show()

    def show_inventory(self):
        # Obtener datos de inventario desde el controlador
        inventory_controller = self.controller.get_inventory()

        # print(inventory_controller)  # Añade esta línea para depurar

        # Limpiar la tabla
        self.table_inventory.setRowCount(0)

        self.table_inventory.verticalHeader().setDefaultSectionSize(200)
        self.table_inventory.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar inventario en la tabla
        for i, inventory in enumerate(inventory_controller):
            self.table_inventory.insertRow(i)

            # Verifica si inventory es una tupla o una lista de un solo valor
            if isinstance(inventory, (tuple, list)):
                # ID
                self.table_inventory.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(inventory[0]))
                )

                # Imagen (asumiendo que está en la columna 1, posición de "name")
                if inventory[1] is not None:  # Asegúrate de que el BLOB no sea nulo
                    pixmap = QPixmap()
                    pixmap.loadFromData(inventory[1])  # Convierte el BLOB en una imagen
                    image_label = QtWidgets.QLabel()  # Crea un QLabel para la imagen
                    image_label.setPixmap(
                        pixmap.scaled(200, 200, QtCore.Qt.KeepAspectRatio)
                    )  # Ajusta el tamaño de la imagen
                    self.table_inventory.setCellWidget(
                        i, 1, image_label
                    )  # Muestra la imagen en la tabla
                else:
                    self.table_inventory.setItem(
                        i, 1, QtWidgets.QTableWidgetItem("No image")
                    )  # Muestra texto si no hay imagen

                # Otros datos
                self.table_inventory.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(inventory[2])
                )  # Name
                self.table_inventory.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(inventory[3])
                )  # Category
                self.table_inventory.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(str(inventory[4]))
                )  # Stock
                self.table_inventory.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(str(inventory[5]))
                )  # Purchase Price
                self.table_inventory.setItem(
                    i, 6, QtWidgets.QTableWidgetItem(str(inventory[6]))
                )  # Sale Price
                self.table_inventory.setItem(
                    i, 7, QtWidgets.QTableWidgetItem(str(inventory[7]))
                )  # Total purch
                self.table_inventory.setItem(
                    i, 8, QtWidgets.QTableWidgetItem(inventory[8])
                )  # Description
                self.table_inventory.setItem(
                    i, 9, QtWidgets.QTableWidgetItem(inventory[9])
                )  # Type Product
            else:
                # Si inventory es un valor único, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {inventory}"
                )

    def update_inventory(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_inventory.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_inventory.item(selected_row, 0)
            # image_item = self.table_inventory.item(selected_row, 1)
            name_item = self.table_inventory.item(selected_row, 2)
            category_item = self.table_inventory.item(selected_row, 3)
            stock_item = self.table_inventory.item(selected_row, 4)
            pricepurch_item = self.table_inventory.item(selected_row, 5)
            pricesale_item = self.table_inventory.item(selected_row, 6)
            totalpurch_item = self.table_inventory.item(selected_row, 7)
            description_item = self.table_inventory.item(selected_row, 8)

            if (
                uid_item
                and name_item
                and category_item
                and stock_item
                and pricepurch_item
                and pricesale_item
                and totalpurch_item
                and description_item
            ):
                uid = uid_item.text()
                name = name_item.text()
                category = category_item.text()
                stock = stock_item.text()
                pricepurch = pricepurch_item.text()
                pricesale = pricesale_item.text()
                totalpurch = totalpurch_item.text()
                description = description_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(
                    uid,
                    name,
                    category,
                    stock,
                    pricepurch,
                    pricesale,
                    totalpurch,
                    description,
                )
                self.update_form.show()
                self.show_inventory()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def export_excel_inventory(self):
        # Obtener datos de usuarios desde el controlador
        inventory_controller = self.controller.get_inventory()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            inventory_controller,
            columns=[
                "ID",
                "Image",
                "Name",
                "Category",
                "Stock",
                "Purchase Price",
                "Sale Price",
                "Total",
                "Description",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Inventario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_inventory.rowCount()):
            match = False
            for col in range(self.table_inventory.columnCount()):
                item = self.table_inventory.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_inventory.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
