import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllergoal import GoalController
from view.admin.goal.viewadd import Viewadd
from view.admin.goal.viewupdate import Viewupdate


class Viewmaingoals(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmaingoals, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/standar/maingoal{self.theme}.ui", self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        icon_configurate_top(self)
        icon_exit_program(self)
        icon_excel(self)

        self.btn_get.clicked.connect(self.show)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = GoalController(self)

        self.show()

    def close_program(self):
        QApplication.quit()

    def show(self):
        # Obtener datos desde el controlador (usando el INNER JOIN)
        Goal_controller = self.controller.get()

        # Limpiar la tabla antes de insertar nuevos datos
        self.table_goal.setRowCount(0)

        # Configuración del tamaño de las filas y columnas
        self.table_goal.verticalHeader().setDefaultSectionSize(200)
        self.table_goal.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar los datos obtenidos en la tabla
        for i, goal in enumerate(Goal_controller):
            self.table_goal.insertRow(i)

            # Verifica si `goal` es una tupla o lista con múltiples elementos
            if isinstance(goal, (tuple, list)):
                # Asumiendo que la consulta INNER JOIN devuelve al menos estos campos:
                # (user_id, user_name, task_id, task_description, task_status)
                self.table_goal.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(goal[0]))
                )  # goal_id
                self.table_goal.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(goal[1])
                )  # name_goal
                self.table_goal.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(goal[2])
                )  # description
                self.table_goal.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(goal[3])
                )  # status
                self.table_goal.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(goal[4])
                )  # start_date
                self.table_goal.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(goal[5])
                )  # end_date
                self.table_goal.setItem(
                    i, 6, QtWidgets.QTableWidgetItem(str(goal[6]))
                )  # id_business
                self.table_goal.setItem(i, 7, QtWidgets.QTableWidgetItem(goal[7]))
            else:
                # Si `goal` no es una tupla o lista, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {goal}"
                )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        goal_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            goal_controller,
            columns=[
                "ID",
                "Nombre",
                "Descripción",
                "Estado",
                "Fecha de inicio",
                "Fecha de termino",
                "ID de la empresa",
                "Nombre de la empresa",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_goal.rowCount()):
            match = False
            for col in range(self.table_goal.columnCount()):
                item = self.table_goal.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_goal.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
