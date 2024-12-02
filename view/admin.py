# ==================imports===================
import webbrowser
import sqlite3
import os
from openpyxl import Workbook
import openpyxl
import matplotlib.pyplot as plt
# bliblioteca para persistencia del tema del software
import configparser
# herramienta para almacenar y iniciar sesión con hash de contraseña
import bcrypt
from passlib.hash import bcrypt
import random
import string
from tkinter import *
import tkinter as tk
#from manager_file import ExploradorArchivos
from tkinter import ttk, messagebox, filedialog, Menu, PhotoImage, scrolledtext
from ttkbootstrap import Style
from time import strftime
from datetime import date
from tkinter import scrolledtext as tkst
from PIL import Image, ImageTk

# ============================================

root = Tk()
root.geometry("1366x768")
root.title("EmprendeSoft(ADMIN)")

user = StringVar()
passwd = StringVar()
fname = StringVar()
lname = StringVar()



def random_emp_id(stringLength):
    Digits = string.digits
    strr = "".join(random.choice(Digits) for i in range(stringLength - 3))
    return "EMP" + strr

class login_page:

    def read_theme_config():
        config = configparser.ConfigParser()
        config.read("config.ini")
        theme_name = config.get("Theme", "name", fallback="default_theme")
        return theme_name

    def change_theme(theme_name):
        # Guardar el tema seleccionado en el archivo de configuración
        config = configparser.ConfigParser()
        config["Theme"] = {"name": theme_name}

        with open("config.ini", "w") as configfile:
            config.write(configfile)

        # Aplicar el nuevo tema
        Style(theme=theme_name)


    def center_window(self, top=None):
        if top is None:
            top = (
                self.top
            )  # Si no se proporciona una ventana específica, utiliza la ventana principal

        # Obtiene el ancho y alto de la pantalla
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()

        # Obtiene el ancho y alto de la ventana
        window_width = top.winfo_reqwidth()
        window_height = top.winfo_reqheight()

        # Calcula las coordenadas x e y para centrar la ventana
        x = (screen_width - window_width) // 4
        y = (screen_height - window_height) // 4

        # Centra la ventana en las coordenadas calculadas
        top.geometry("+{}+{}".format(x, y))
    # vista login
    def __init__(self, top=None):
        self.top = top
        self.top.geometry("800x500")
        self.top.resizable(0, 0)
        self.top.title("EmprendeSoft - Inicio de sesión (Administrador)")
        self.top.iconbitmap("img/icon.ico")
          # Leer y aplicar el tema guardado previamente al iniciar la aplicación
        initial_theme = read_theme_config()
        change_theme(initial_theme)
        # frame_logo
        frame_logo = tk.Frame(
            self.top, bd=0, width=300, relief=tk.SOLID, padx=10, pady=10, bg="#BB2B2B"
        )
        frame_logo.pack(side="left", expand=tk.YES, fill=tk.BOTH)
        self.label_logo = tk.Label(frame_logo, bg="#BB2B2B")
        self.label_logo.place(x=0, y=0, relwidth=1, relheight=1)
        self.cargar_logo_guardado()
        icono = tk.PhotoImage(file="img/perfil_git.gif")
        self.top.iconphoto(True, icono)
        self.center_window()

        # frame_form
        frame_form = tk.Frame(top, bd=0, relief=tk.SOLID, bg="#fcfcfc")
        frame_form.pack(side="right", expand=tk.YES, fill=tk.BOTH)
        # frame_form

        # frame_form_top
        frame_form_top = tk.Frame(
            frame_form, height=50, bd=0, relief=tk.SOLID, bg="black"
        )
        frame_form_top.pack(side="top", fill=tk.X)
        title = tk.Label(
            frame_form_top,
            text="Iniciar Sesión",
            font=("Times", 30),
            fg="#666a88",
            bg="#fcfcfc",
            pady=50,
        )
        title.pack(expand=tk.YES, fill=tk.BOTH)
        # end frame_form_top

        # frame_form_fill
        frame_form_fill = tk.Frame(
            frame_form, height=50, bd=0, relief=tk.SOLID, bg="#fcfcfc"
        )
        frame_form_fill.pack(side="bottom", expand=tk.YES, fill=tk.BOTH)

        # Etiqueta para usuario
        self.user_label = Label(
            root, text="Usuario", font=("Times", 15), foreground="#000000"
        )
        self.user_label.place(relx=0.6, rely=0.30)

        self.entry1 = Entry(root)
        self.entry1.place(relx=0.6, rely=0.35, width=300, height=24)
        self.entry1.configure(font="-family {Poppins} -size 10")
        self.entry1.configure(textvariable=user)

        # Etiqueta para contraseña
        self.user_label = Label(
            root, text="Contraseña", font=("Times", 15), foreground="#000000"
        )
        self.user_label.place(relx=0.6, rely=0.45)

        self.entry2 = Entry(root)
        self.entry2.place(relx=0.6, rely=0.50, width=300, height=24)
        self.entry2.configure(font="-family {Poppins} -size 10")
        self.entry2.configure(show="*")
        self.entry2.configure(textvariable=passwd)

        # Cargar el icono
        self.icono_login = Image.open("img/icon.ico")
        self.icono_login = self.icono_login.resize(
            (32, 32), Image.LANCZOS
        )  # Utiliza LANCZOS en lugar de ANTIALIAS
        self.icono_login = ImageTk.PhotoImage(self.icono_login)
        top.iconphoto(True, self.icono_login)

        self.button1 = Button(root)
        self.button1.place(relx=0.6, rely=0.685, width=300, height=43)
        self.button1.configure(relief="flat")
        self.button1.configure(overrelief="flat")
        self.button1.configure(activebackground="#3a7ff6")
        self.button1.configure(cursor="hand2")
        self.button1.configure(foreground="#ffffff")
        self.button1.configure(background="#BB2B2B")
        self.button1.configure(
            font="-family {Times SemiBold} -size 20", borderwidth="0", text="Iniciar"
        )
        self.button1.configure(command=self.login)

    def abrir_logo(self):
        ruta_logo = filedialog.askopenfilename(
            title="Seleccionar logo",
            filetypes=[("Archivos de imagen", "*.png;*.jpg;*.jpeg")],
        )
        if ruta_logo:
            self.guardar_ruta_logo(ruta_logo)
            self.cargar_logo(ruta_logo)

    def guardar_ruta_logo(self, ruta_logo):
        with open("config.txt", "w") as f:
            f.write(ruta_logo)

    def cargar_logo_guardado(self):
        if os.path.exists("config.txt"):
            with open("config.txt", "r") as f:
                ruta_logo = f.read()
                if os.path.exists(ruta_logo):
                    self.cargar_logo(ruta_logo)

    def cargar_logo(self, ruta_logo):
        imagen = Image.open(ruta_logo)
        imagen = imagen.resize((150, 150), resample=Image.LANCZOS)
        foto = ImageTk.PhotoImage(imagen)
        self.label_logo.config(image=foto)
        self.label_logo.image = foto

    # controlador del login
    def login(self, Event=None):
        global username
        username = user.get()
        password = passwd.get()

        find_user = "SELECT * FROM employee WHERE name = ?"
        cur.execute(find_user, [username])
        user_data = cur.fetchone()  # Obtenemos los datos del usuario

        if user_data:
            # Obtenemos el hash de la contraseña almacenada en la base de datos
            stored_password_hash = user_data[5]  # Suponiendo que el hash está en la sexta columna
            # Verificamos si la contraseña ingresada coincide con el hash almacenado
            if bcrypt.verify(password, stored_password_hash):
                if user_data[6] == "administrador":
                    messagebox.showinfo("Mensaje", "Correcto, bienvenido.")
                    page1.entry1.delete(0, END)
                    page1.entry2.delete(0, END)

                    root.withdraw()
                    global adm
                    global page2
                    adm = Toplevel()
                    page2 = Admin_Page(adm)
                    # page2.time()
                    adm.protocol("WM_DELETE_WINDOW", exitt)
                    adm.mainloop()
                    db.close()
                else:
                    messagebox.showerror(
                        "mensaje de error", "Usted no es administrador."
                    )
            else:
                messagebox.showerror("Error", "Contraseña incorrecta.")
                passwd.delete(0, END)  # Limpiar el campo de contraseña
        else:
            messagebox.showerror("Error", "Cuenta incorrecta.")
            passwd.delete(0, END)  # Limpiar el campo de contraseña     

def exitt():
    sure = messagebox.askyesno("Salir", "Está se guro de salir.", parent=root)
    if sure == True:
        adm.destroy()
        root.destroy()

def inventory():
    # adm.withdraw()
    global inv
    global page3
    inv = Toplevel()
    page3 = Inventory(inv)
    page3.time()
    inv.protocol("WM_DELETE_WINDOW", exitt)
    inv.mainloop()

def employee():
    # adm.withdraw()
    global emp
    global page5
    emp = Toplevel()
    page5 = Employee(emp)
    page5.time()
    emp.protocol("WM_DELETE_WINDOW", exitt)
    emp.mainloop()

def invoices():
    # adm.withdraw()
    global invoice
    invoice = Toplevel()
    page7 = Invoice(invoice)
    page7.time()
    invoice.protocol("WM_DELETE_WINDOW", exitt)
    invoice.mainloop()

def clients():
    # adm.withdraw()
    global client
    global page8
    client = Toplevel()
    page8 = Client(client)
    page8.time()
    client.protocol("WM_DELETE_WINDOW", exitt)
    client.mainloop()

def quote():
    # Obtener datos de la base de datos
    cur.execute("SELECT product_name, total_price FROM raw_inventory")
    filas = cur.fetchall()

    # Separar las categorías y los valores en listas
    client = [fila[0] for fila in filas]
    total = [fila[1] for fila in filas]

    # Crear gráfico de barras
    plt.bar(client, total, color="lightgreen")

    # Agregar etiquetas y título
    plt.xlabel("Nombres de los Productos")
    plt.ylabel("Gastos")
    plt.title("Gráfico de gastos")

    # Mostrar gráfico
    plt.show()

def about():
    pass

def graphi_incomi():
    # Obtener datos de la base de datos
    cur.execute("SELECT customer_name, total FROM bill")
    filas = cur.fetchall()

    # Separar las categorías y los valores en listas
    client = [fila[0] for fila in filas]
    total = [fila[1] for fila in filas]

    # Crear gráfico de barras
    plt.bar(client, total, color="lightgreen")

    # Agregar etiquetas y título
    plt.xlabel("Nombres de los Clientes")
    plt.ylabel("Ingresos")
    plt.title("Gráfico de ingresos")

    # Mostrar gráfico
    plt.show()

def graphi_trans():
    # Obtener datos de la base de datos
    cur.execute("SELECT date_transaction, amount FROM transactions")
    filas = cur.fetchall()

    # Separar las categorías y los valores en listas
    client = [fila[0] for fila in filas]
    total = [fila[1] for fila in filas]

    # Crear gráfico de barras
    plt.bar(client, total, color="lightgreen")

    # Agregar etiquetas y título
    plt.xlabel("Fecha de transacción")
    plt.ylabel("Monto")
    plt.title("Gráfico de montos")

    # Mostrar gráfico
    plt.show()

def graphi_inve():
       # Obtener datos de la base de datos
    cur.execute("SELECT type, yield FROM investments")
    filas = cur.fetchall()

    # Separar las categorías y los valores en listas
    client = [fila[0] for fila in filas]
    total = [fila[1] for fila in filas]

    # Crear gráfico de barras
    plt.bar(client, total, color="lightgreen")

    # Agregar etiquetas y título
    plt.xlabel("Tipo de inversión")
    plt.ylabel("Monto")
    plt.title("Gráfico de montos")

    # Mostrar gráfico
    plt.show()

# Función para cambiar el tema
def change_theme(theme_name):
    # Guardar el tema seleccionado en el archivo de configuración
    config = configparser.ConfigParser()
    config["Theme"] = {"name": theme_name}

    with open("config.ini", "w") as configfile:
        config.write(configfile)

    # Aplicar el nuevo tema
    Style(theme=theme_name)

# Función para abrir el archivo de ayuda en PDF
def open_help(file_path):
    # Abre el archivo PDF en el navegador predeterminado
    webbrowser.open(file_path)

    # Leer la configuración actual del archivo de configuración

def read_theme_config():
    config = configparser.ConfigParser()
    config.read("config.ini")
    theme_name = config.get("Theme", "name", fallback="default_theme")
    return theme_name

# style = Style(theme="vapor")  # Selecciona el tema que prefieras

def managervoucher():
    # adm.withdraw()
    global vou
    global pagevou
    vou = tk.Toplevel()
    pagevou = ExploradorArchivos(vou)
    pagevou.time()
    vou.protocol("WM_DELETE_WINDOW", exitt)
    vou.mainloop()

def provider():
    # adm.withdraw()
    global provi
    global page3
    provi = tk.Toplevel()
    page3 = Provider(provi)
    page3.time()
    provi.protocol("WM_DELETE_WINDOW", exitt)
    provi.mainloop()
    page3.tree.mainloop()

def transaction():
    # adm.withdraw()
    global trans
    global page_transaction
    trans = Toplevel()
    page_transaction = Transaction(trans)
    page_transaction.time()
    trans.protocol("WM_DELETE_WINDOW", exitt)
    trans.mainloop()

def project():
    # adm.withdraw()
    global pro
    global page_project
    pro = Toplevel()
    page_project = Project(pro)
    page_project.time()
    pro.protocol("WM_DELETE_WINDOW", exitt)
    pro.mainloop()

def sent():
    # adm.withdraw()
    global se
    global page_sent
    se = Toplevel()
    page_sent = Sent(se)
    page_sent.time()
    se.protocol("WM_DELETE_WINDOW", exitt)
    se.mainloop()

def investment():
    # adm.withdraw()
    global inves
    global page_investment
    inves = Toplevel()
    page_investment = Investment(inves)
    page_investment.time()
    inves.protocol("WM_DELETE_WINDOW", exitt)
    inves.mainloop()

def task():
    # adm.withdraw()
    global taskk
    global page_task
    taskk = Toplevel()
    page_task = Task(taskk)
    page_task.time()
    taskk.protocol("WM_DELETE_WINDOW", exitt)
    taskk.mainloop()

def goal():
    # adm.withdraw()
    global goals
    global page_goal
    goals = Toplevel()
    page_goal = Goal(goals)
    page_goal.time()
    goals.protocol("WM_DELETE_WINDOW", exitt)
    goals.mainloop()

def upload_logo(self):
        if os.path.exists("config.txt"):
            with open("config.txt", "r") as f:
                ruta_logo = f.read()
                if os.path.exists(ruta_logo):
                    self.cargar_logo(ruta_logo)

class Admin_Page:
    def __init__(self, top=None):
        # Agregar opciones de temas al menú
        self.menu_bar = Menu(top)
        top.config(menu=self.menu_bar)

        theme_menu = Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Configuración de Temas", menu=theme_menu)

        theme_menu.add_command(label="--Temas Claros--")
        theme_menu.add_command(
            label="Cerculean", command=lambda: change_theme("cerculean")
        )
        theme_menu.add_command(label="Simplex", command=lambda: change_theme("simplex"))
        theme_menu.add_command(label="Morph", command=lambda: change_theme("morph"))
        theme_menu.add_command(label="Yeti", command=lambda: change_theme("yeti"))
        theme_menu.add_command(label="United", command=lambda: change_theme("united"))
        theme_menu.add_command(
            label="Sandstone", command=lambda: change_theme("sandstone")
        )
        theme_menu.add_command(label="Pulse", command=lambda: change_theme("pulse"))
        theme_menu.add_command(label="Minty", command=lambda: change_theme("minty"))
        theme_menu.add_command(label="Lumen", command=lambda: change_theme("lumen"))
        theme_menu.add_command(label="Litera", command=lambda: change_theme("litera"))
        theme_menu.add_command(label="Journal", command=lambda: change_theme("journal"))
        theme_menu.add_command(label="Flatly", command=lambda: change_theme("flatly"))
        theme_menu.add_command(label="Cosmo", command=lambda: change_theme("cosmo"))
        theme_menu.add_command(label="--Temas Oscuros--")
        theme_menu.add_command(label="Solar", command=lambda: change_theme("solar"))
        theme_menu.add_command(
            label="Superhero", command=lambda: change_theme("superhero")
        )
        theme_menu.add_command(label="Oscuro", command=lambda: change_theme("darkly"))
        theme_menu.add_command(label="Cyborg", command=lambda: change_theme("cyborg"))
        theme_menu.add_command(label="Vapor", command=lambda: change_theme("vapor"))

        # Leer y aplicar el tema guardado previamente al iniciar la aplicación
        initial_theme = read_theme_config()
        change_theme(initial_theme)

        # Agregar menú de Ayuda con subcategorías
        help_menu = Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Ayuda", menu=help_menu)

        # Agregar opciones de ayuda con redirección a archivos PDF
        help_menu.add_command(
            label="Guía de Usuario", command=lambda: open_help("hoja.pdf")
        )
        help_menu.add_command(
            label="Manual del Administrador",
            command=lambda: open_help("ruta/al/manual_del_administrador.pdf"),
        )

        # Agrega más opciones de temas según tus necesidades
        # Agregar opciones de temas al menú
        top.geometry("1200x700")
        top.state("zoomed")
        top.title("EmprendeSoft - Panel del Administrador")
        self.top = top
        # está función es para maximizar la ventana

        self.message = Label(top)
        self.message.place(relx=0.855, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")
        
        self.clock = Label(top)
        self.clock.place(relx=0.5, rely=0.1, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        icon_path_exit = "img/salida.png"

        self.icon_exit = PhotoImage(file=icon_path_exit)

        self.button1 = ttk.Button(
            top,
            text="Cerrar Sesión",
            style="danger.TButton",
            image=self.icon_exit,
            command=self.Logout,
        )
        self.button1.place(relx=0.855, rely=0.106, width=110, height=35)
        self.button1.configure(cursor="hand2")

        # Agrega la ruta de tu archivo de icono
        icon_path_inventario = "img/proveedor-alternativo.png"

        # Agrega la ruta de tu archivo de icono
        icon_path_empleado = "img/hombre-empleado.png"

        # Agrega la ruta de tu archivo de icono
        icon_path_factura = "img/archivo-factura-dolar.png"

        # Agrega la ruta de tu archivo de icono
        icon_path_client = "img/icon-client.png"

        # Agrega la ruta de tu archivo icono
        icon_path_graphi = "img/grafico.png"

        # Agregar la ruta de tu archivi icono
        icon_path_quote = "img/decreciente.png"

        # Agregar la ruta de tu archivo
        icon_path_pdf = "img/pdf.png"

        # Agregar la ruta de tu archivo icono
        icon_path_distribution = "img/distribution.png"

        icon_path_transaction = "img/card-payment.png"

        icon_path_graph_transaction = "img/grafic-trans.png"

        icon_path_project = "img/gestion-de-proyectos.png"

        icon_path_sent = "img/camion-de-reparto.png"

        icon_path_investment = "img/investment.png"

        icon_path_ganancia = "img/ganancias.png"

        icon_path_tarea = "img/tareas.png"

        icon_path_meta = "img/metas.png"

        self.icon_inventario = PhotoImage(file=icon_path_inventario)
        self.icon_empleado = PhotoImage(file=icon_path_empleado)
        self.icon_factura = PhotoImage(file=icon_path_factura)
        self.icon_client = PhotoImage(file=icon_path_client)
        self.icon_grafi = PhotoImage(file=icon_path_graphi)
        self.icon_quote = PhotoImage(file=icon_path_quote)
        self.icon_pdf = PhotoImage(file=icon_path_pdf)
        self.icon_distribution = PhotoImage(file=icon_path_distribution)
        self.icon_trnsaction = PhotoImage(file=icon_path_transaction)
        self.icon_grapi_transaction = PhotoImage(file=icon_path_graph_transaction)
        self.icon_project = PhotoImage(file=icon_path_project)
        self.icon_sent = PhotoImage(file=icon_path_sent)
        self.icon_investment = PhotoImage(file=icon_path_investment)
        self.icon_ganancias = PhotoImage(file=icon_path_ganancia)
        self.icon_metas = PhotoImage(file=icon_path_meta)
        self.icon_tareas = PhotoImage(file=icon_path_tarea)

        self.button_inventory = ttk.Button(
            top,
            text="Inventario",
            style="info.TButton",
            command=inventory,
            compound="top",
            image=self.icon_inventario,
        )
        self.button_inventory.place(relx=0.14, rely=0.1, width=120, height=100)
        self.button_inventory.configure(cursor="hand2")

        self.button_employee = ttk.Button(
            top,
            text="Empleados",
            style="success.TButton",
            command=employee,
            compound="top",
            image=self.icon_empleado,
        )
        self.button_employee.place(relx=0.338, rely=0.1, width=120, height=100)
        self.button_employee.configure(cursor="hand2")

        self.button_invoice = ttk.Button(
            top,
            text="Facturas",
            style="info.TButton",
            command=invoices,
            compound="top",
            image=self.icon_factura,
        )
        self.button_invoice.place(relx=0.536, rely=0.1, width=120, height=100)
        self.button_invoice.configure(cursor="hand2")

        self.button_client = ttk.Button(
            top,
            text="Clientes",
            style="success.TButton",
            command=clients,
            compound="top",
            image=self.icon_client,
        )
        self.button_client.place(relx=0.730, rely=0.1, width=120, height=100)
        self.button_client.configure(cursor="hand2")

        self.button_graphic = ttk.Button(
            top,
            text="Gráfico de ingresos",
            style="info.TButton",
            command=graphi_incomi,
            compound="top",
            image=self.icon_grafi,
        )
        self.button_graphic.place(relx=0.14, rely=0.3, width=120, height=100)
        self.button_graphic.configure(cursor="hand2")

        self.button_quote = ttk.Button(
            top,
            text="Gastos",
            style="success.TButton",
            command=quote,
            compound="top",
            image=self.icon_quote,
        )
        self.button_quote.place(relx=0.338, rely=0.3, width=120, height=100)
        self.button_quote.configure(cursor="hand2")

        self.button_manager_pdf = ttk.Button(
            top,
            text="Boleta PDF",
            style="info.TButton",
            command=managervoucher,
            compound="top",
            image=self.icon_pdf,
        )
        self.button_manager_pdf.place(relx=0.536, rely=0.3, width=120, height=100)
        self.button_manager_pdf.configure(cursor="hand2")

        self.button_provider = ttk.Button(
            top,
            text="Proveedor",
            style="success.TButton",
            command=provider,
            compound="top",
            image=self.icon_distribution,
        )
        self.button_provider.place(relx=0.730, rely=0.3, width=120, height=100)
        self.button_provider.configure(cursor="hand2")

        self.button_transaction = ttk.Button(
            top,
            text="Transacciones",
            style="info.TButton",
            command=transaction,
            compound="top",
            image=self.icon_trnsaction,
        )
        self.button_transaction.place(relx=0.14, rely=0.5, width=120, height=100)
        self.button_transaction.configure(cursor="hand2")

        self.button_graphi_transaction = ttk.Button(
            top,
            text="Gráfico Transacción",
            style="success.TButton",
            command=graphi_trans,
            compound="top",
            image=self.icon_grapi_transaction,
        )
        self.button_graphi_transaction.place(relx=0.338, rely=0.5, width=120, height=100)
        self.button_graphi_transaction.configure(cursor="hand2")

        self.button_project = ttk.Button(
            top,
            text="Proyectos",
            style="info.TButton",
            command=project,
            compound="top",
            image=self.icon_project,
        )
        self.button_project.place(relx=0.536, rely=0.5, width=120, height=100)
        self.button_project.configure(cursor="hand2")

        self.button_send = ttk.Button(
            top,
            text="Envíos",
            style="success.TButton",
            command=sent,
            compound="top",
            image=self.icon_sent,
        )
        self.button_send.place(relx=0.730, rely=0.5, width=120, height=100)
        self.button_send.configure(cursor="hand2")

        self.button_investments = ttk.Button(
            top,
            text="Inversiones",
            style="info.TButton",
            command=investment,
            compound="top",
            image=self.icon_investment,
        )
        self.button_investments.place(relx=0.14, rely=0.7, width=120, height=100)
        self.button_investments.configure(cursor="hand2")

        self.button_investments_graphi = ttk.Button(
            top,
            text="G. Inversiones",
            style="success.TButton",
            command=graphi_inve,
            compound="top",
            image=self.icon_ganancias,
        )
        self.button_investments_graphi.place(relx=0.338, rely=0.7, width=120, height=100)
        self.button_investments_graphi.configure(cursor="hand2")

        self.button_task = ttk.Button(
            top,
            text="Tareas",
            style="info.TButton",
            command=task,
            compound="top",
            image=self.icon_tareas,
        )
        self.button_task.place(relx=0.536, rely=0.7, width=120, height=100)
        self.button_task.configure(cursor="hand2")

        self.button_goals = ttk.Button(
            top,
            text="Metas",
            style="success.TButton",
            command=goal,
            compound="top",
            image=self.icon_metas,
        )
        self.button_goals.place(relx=0.730, rely=0.7, width=120, height=100)
        self.button_goals.configure(cursor="hand2")
        

        # Agrega la ruta de tu archivo de icono principal
        icon_path_principal = "img/icon.png"

        self.icon_principal = PhotoImage(file=icon_path_principal)

        # Crear un botón para el icono principal
        self.button_principal = ttk.Button(
            top,
            text="EmprendeSoft",
            style="default-link.TButton",
            compound="top",
            image=self.icon_principal,
        )
        self.button_principal.place(relx=0.035, rely=0.050, width=100, height=100)

    def Logout(self):
        sure = messagebox.askyesno(
            "Salir", "¿Está seguro de cerrar la sesión", parent=adm
        )
        if sure == True:
            adm.destroy()
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

class Inventory:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Inventario")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(inv)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(inv)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_product,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)


        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Producto",
            style="info-outline.TButton",
            command=self.add_product,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Producto",
            style="success-outline.TButton",
            command=self.update_product,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Producto",
            style="danger-outline.TButton",
            command=self.delete_product,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(inv, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(inv, orient=VERTICAL)
        self.tree = ttk.Treeview(inv)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "Product ID",
                "Name",
                "Category",
                "Sub-Category",
                "In Stock",
                "MRP",
                "Cost Price",
                "Vendor No.",
                "Total",
            )
        )

        self.tree.heading("Product ID", text="ID Producto", anchor=W)
        self.tree.heading("Name", text="Nombre", anchor=W)
        self.tree.heading("Category", text="Categoría", anchor=W)
        self.tree.heading("Sub-Category", text="Sub-Categoría", anchor=W)
        self.tree.heading("In Stock", text="Cantidad disponible", anchor=W)
        self.tree.heading("MRP", text="MRP", anchor=W)
        self.tree.heading("Cost Price", text="Precio", anchor=W)
        self.tree.heading("Vendor No.", text="Número del vendedor.", anchor=W)
        self.tree.heading("Total", text="Total", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)
        self.tree.column("#7", stretch=NO, minwidth=0, width=80)
        self.tree.column("#8", stretch=NO, minwidth=0, width=100)
        self.tree.column("#9", stretch=NO, minwidth=0, width=100)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM raw_inventory")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_product(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id del producto invalido.", parent=inv)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID Producto: {} encontrado.".format(self.entry1.get()),
                        parent=inv,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} no encontrado.".format(self.entry1.get()),
                    parent=inv,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_product(self):
        val = []
        to_delete = []

        if len(self.sel) != 0:
            sure = messagebox.askyesno(
                "Confirmar", "¿Está seguro de eliminar producto?", parent=inv
            )
            if sure == True:
                for i in self.sel:
                    for j in self.tree.item(i)["values"]:
                        val.append(j)

                for j in range(len(val)):
                    if j % 8 == 0:
                        to_delete.append(val[j])

                for k in to_delete:
                    delete = "DELETE FROM raw_inventory WHERE product_id = ?"
                    cur.execute(delete, [k])
                    db.commit()

                messagebox.showinfo("Mensaje", "Producto eliminado.", parent=inv)
                self.sel.clear()
                self.tree.delete(*self.tree.get_children())

                self.DisplayData()
        else:
            messagebox.showerror("Error", "Por favor seleccionar producto.", parent=inv)

    def update_product(self):
        if len(self.sel) == 1:
            global p_update
            p_update = Toplevel()
            page9 = Update_Product(p_update)
            page9.time()
            p_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page9.entry1.insert(0, valll[1])
            page9.entry2.insert(0, valll[2])
            page9.entry3.insert(0, valll[4])
            page9.entry4.insert(0, valll[5])
            page9.entry6.insert(0, valll[3])
            page9.entry7.insert(0, valll[6])
            page9.entry8.insert(0, valll[7])
            page9.entry_total.insert(0, valll[8])

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=inv
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=inv
            )

        p_update.mainloop()

    def add_product(self):
        global p_add
        global page4
        p_add = Toplevel()
        page4 = add_product(p_add)
        page4.time()
        p_add.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=inv)
        if sure == True:
            inv.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Información", "¿Está seguro de volver")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM raw_inventory"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/inventario.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_product:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Producto")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(p_add)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del Producto", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para nombre del producto
        self.label_nombre = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(p_add)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_category_product = Label(
            top, text="Categoría del producto", font=("Times", 15), foreground="#fff"
        )
        self.label_category_product.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(p_add)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        self.r2 = p_add.register(self.testint)

        # Etiqueta para stock del producto
        self.label_stock_product = Label(
            top, text="Cantidad del producto", font=("Times", 15), foreground="#fff"
        )

        self.label_stock_product.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(p_add)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para mrp
        self.label_mrp = Label(
            top, text="Precio Venta", font=("Times", 15), foreground="#fff"
        )

        self.label_mrp.place(relx=0.132, rely=0.590)

        self.entry4 = Entry(p_add)
        self.entry4.place(relx=0.132, rely=0.646, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_sub_category = Label(
            top, text="Sub categoría", font=("Times", 15), foreground="#fff"
        )

        self.label_sub_category.place(relx=0.527, rely=0.370)

        self.entry6 = Entry(p_add)
        self.entry6.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_price_product = Label(
            top, text="Precio Compra", font=("Times", 15), foreground="#fff"
        )

        self.label_price_product.place(relx=0.527, rely=0.480)

        self.entry7 = Entry(p_add)
        self.entry7.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry7.configure(font="-family {Poppins} -size 12")

        # Etiqueta para phone seller
        self.label_phone_seller = Label(
            top, text="Teléfono del vendedor", font=("Times", 15), foreground="#fff"
        )

        self.label_phone_seller.place(relx=0.527, rely=0.599)

        self.entry8 = Entry(p_add)
        self.entry8.place(relx=0.527, rely=0.646, width=374, height=30)
        self.entry8.configure(font="-family {Poppins} -size 12")
        self.entry8.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para total del precio
        self.label_total = Label(
            top, text="Total", font=("Times", 15), foreground="#fff"
        )
        self.label_total.place(relx=0.132, rely=0.700)

        self.entry_total = Entry(p_add)
        self.entry_total.place(relx=0.132, rely=0.760, width=730, height=30)
        self.entry_total.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

        self.button_calcular_total = ttk.Button(
            top,
            text="Calcular total",
            style="danger.TButton",
            command=self.calculartotal,
        )
        self.button_calcular_total.place(relx=0.640, rely=0.836, width=100, height=34)
        self.button_calcular_total.configure(cursor="hand2")

    def add(self):
        pqty = self.entry3.get()
        pcat = self.entry2.get()
        pmrp = self.entry4.get()
        pname = self.entry1.get()
        psubcat = self.entry6.get()
        pcp = self.entry7.get()
        pvendor = self.entry8.get()
        total = self.entry_total.get()

        if pname.strip():
            if pcat.strip():
                if psubcat.strip():
                    if pqty:
                        if pcp:
                            try:
                                float(pcp)
                            except ValueError:
                                messagebox.showerror(
                                    "Error",
                                    "Invalido precio venta del producto",
                                    parent=p_add,
                                )
                            else:
                                if pmrp:
                                    try:
                                        float(pmrp)
                                    except ValueError:
                                        messagebox.showerror(
                                            "Error",
                                            "Precio compra invalido",
                                            parent=p_add,
                                        )
                                    if total.strip():
                                        # Realiza operaciones en la base de datos aquí
                                        # ...
                                        insert = "INSERT INTO raw_inventory(product_name, product_cat, product_subcat, stock, mrp, cost_price, vendor_phn, total_price) VALUES(?,?,?,?,?,?,?,?)"
                                        cur.execute(
                                        insert,
                                        [
                                        pname,
                                        pcat,
                                        psubcat,
                                        int(pqty),
                                        float(pmrp),
                                        float(pcp),
                                        pvendor,
                                        total,
                                        ],
                                        )
                                        db.commit()
                                        messagebox.showinfo(
                                        "Información",
                                        "Producto ingresado correctamente",
                                        parent=p_add,
                                        )
                                        p_add.destroy()
                                        page3.tree.delete(*page3.tree.get_children())
                                        page3.DisplayData()
                                        p_add.destroy()
                                    else:
                                        messagebox.showerror(
                                            "Error",
                                            "Debe calcular el total",
                                            parent=p_add,
                                        )
                                else:
                                    messagebox.showerror(
                                        "Error",
                                        "Por favor ingrese precio compra del producto ",
                                        parent=p_add,
                                    )
                        else:
                            messagebox.showerror(
                                "Error",
                                "Por favor ingresar el precio venta del producto",
                                parent=p_add,
                            )
                    else:
                        messagebox.showerror(
                            "Error",
                            "Por favor ingrese cantidad de producto disponibles",
                            parent=p_add,
                        )
                else:
                    messagebox.showerror(
                        "Error", "Por favor ingrese categoría", parent=p_add
                    )
            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=p_add
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=p_add)
        

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)
        self.entry7.delete(0, END)
        self.entry8.delete(0, END)

    def calculartotal(self):
        try:
            # Obtiene el precio de venta ingresado por el usuario
            precio_venta = float(self.entry4.get())
            # Obtiene la cantidad del producto ingresada por el usuario
            cantidad_producto = float(self.entry3.get())
            # Calcula el total multiplicando el precio de venta por la cantidad del producto
            total = precio_venta * cantidad_producto
            # Muestra el total en el campo de entrada correspondiente
            self.entry_total.delete(
                0, END
            )  # Borra cualquier valor anterior en el campo de entrada
            self.entry_total.insert(0, str(total))  # Inserta el nuevo total calculado
        except ValueError:
            # Maneja el caso en el que los valores ingresados no sean números
            messagebox.showerror(
                "Error",
                "Ingrese números válidos para el precio de venta y la cantidad del producto.",
            )

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Product:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Producto")

        self.clock = Label(p_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Producto", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para nombre del producto
        self.label_nombre = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(p_update)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para categoría del producto
        self.label_category_product = Label(
            top, text="Categoría del producto", font=("Times", 15), foreground="#fff"
        )
        self.label_category_product.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(p_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")

        self.r2 = p_update.register(self.testint)

        # Etiqueta para stock del producto
        self.label_stock_product = Label(
            top, text="Cantidad de productos", font=("Times", 15), foreground="#fff"
        )

        self.label_stock_product.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(p_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        self.entry3.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para mrp
        self.label_mrp = Label(
            top, text="Precio Venta", font=("Times", 15), foreground="#fff"
        )

        self.label_mrp.place(relx=0.132, rely=0.590)

        self.entry4 = Entry(p_update)
        self.entry4.place(relx=0.132, rely=0.646, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")

        # Etiqueta para sub category
        self.label_sub_category = Label(
            top, text="Sub categoría", font=("Times", 15), foreground="#fff"
        )

        self.label_sub_category.place(relx=0.527, rely=0.370)

        self.entry6 = Entry(p_update)
        self.entry6.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        self.entry6.configure(relief="flat")

        # Etiqueta para sub category
        self.label_price_product = Label(
            top, text="Precio Compra", font=("Times", 15), foreground="#fff"
        )

        self.label_price_product.place(relx=0.527, rely=0.480)

        self.entry7 = Entry(p_update)
        self.entry7.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry7.configure(font="-family {Poppins} -size 12")
        self.entry7.configure(relief="flat")

        # Etiqueta para phone seller
        self.label_phone_seller = Label(
            top, text="Teléfono del vendedor", font=("Times", 15), foreground="#fff"
        )

        self.label_phone_seller.place(relx=0.527, rely=0.599)

        self.entry8 = Entry(p_update)
        self.entry8.place(relx=0.527, rely=0.646, width=374, height=30)
        self.entry8.configure(font="-family {Poppins} -size 12")
        self.entry8.configure(relief="flat")

        # Etiqueta para total del precio
        self.label_total = Label(
            top, text="Total", font=("Times", 15), foreground="#fff"
        )
        self.label_total.place(relx=0.132, rely=0.700)

        self.entry_total = Entry(p_update)
        self.entry_total.place(relx=0.132, rely=0.760, width=730, height=30)
        self.entry_total.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Actualizar", style="warning-outline.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top,
            text="Calcular Total",
            style="danger-outline.TButton",
            command=self.calculartotalupdate,
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def update(self):
        pqty = self.entry3.get()
        pcat = self.entry2.get()
        pmrp = self.entry4.get()
        pname = self.entry1.get()
        psubcat = self.entry6.get()
        pcp = self.entry7.get()
        pvendor = self.entry8.get()
        total = self.entry_total.get()

        if pname.strip():
            if pcat.strip():
                if psubcat.strip():
                    if pqty:
                        if pcp:
                            try:
                                float(pcp)
                            except ValueError:
                                messagebox.showerror(
                                    "Error", "Precio venta invalido", parent=p_update
                                )
                            else:
                                if pmrp:
                                    try:
                                        float(pmrp)
                                    except ValueError:
                                        messagebox.showerror(
                                            "Error",
                                            "Precio compra invalido",
                                            parent=p_update,
                                        )
                                    else:

                                        product_id = valll[0]
                                        with connect_to_database() as db:
                                            try:
                                                cur = db.cursor()
                                                update = "UPDATE raw_inventory SET product_name = ?, product_cat = ?, product_subcat = ?, stock = ?, mrp = ?, cost_price = ?, vendor_phn = ?, total_price = ?  WHERE product_id = ?"
                                                cur.execute(
                                                update,
                                                [
                                                pname,
                                                pcat,
                                                psubcat,
                                                int(pqty),
                                                float(pmrp),
                                                float(pcp),
                                                pvendor,
                                                total,
                                                product_id,
                                                ],
                                                )
                                                db.commit()
                                                messagebox.showinfo(
                                                "Información",
                                                "Producto actualizado correctamente",
                                                parent=p_update,
                                                )
                                                valll.clear()
                                                Inventory.sel.clear()
                                                page3.tree.delete(*page3.tree.get_children())
                                                page3.DisplayData()
                                                p_update.destroy()
                                            except sqlite3.Error as e:
                                                print("Error en la consulta:", e)
                                            finally:
                                                db.close()  # Cierra la conexión al finalizar

                                else:
                                    messagebox.showerror(
                                        "Error",
                                        "Por favor agregar precio de compra del producto",
                                        parent=p_update,
                                    )
                        else:
                            messagebox.showerror(
                                "Oops!",
                                "Please enter product cost price.",
                                parent=p_update,
                            )
                    else:
                        messagebox.showerror(
                            "Oops!", "Please enter product quantity.", parent=p_update
                        )
                else:
                    messagebox.showerror(
                        "Oops!", "Please enter product sub-category.", parent=p_update
                    )
            else:
                messagebox.showerror(
                    "Oops!", "Please enter product category.", parent=p_update
                )
        else:
            messagebox.showerror("Oops!", "Please enter product name", parent=p_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)
        self.entry7.delete(0, END)
        self.entry8.delete(0, END)

    def calculartotalupdate(self):
        try:
            # Obtiene el precio de venta ingresado por el usuario
            precio_venta = float(self.entry4.get())
            # Obtiene la cantidad del producto ingresada por el usuario
            cantidad_producto = float(self.entry3.get())
            # Calcula el total multiplicando el precio de venta por la cantidad del producto
            total = precio_venta * cantidad_producto
            # Muestra el total en el campo de entrada correspondiente
            self.entry_total.delete(
                0, END
            )  # Borra cualquier valor anterior en el campo de entrada
            self.entry_total.insert(0, str(total))  # Inserta el nuevo total calculado
        except ValueError:
            # Maneja el caso en el que los valores ingresados no sean números
            messagebox.showerror(
                "Error",
                "Ingrese números válidos para el precio de venta y la cantidad del producto.",
            )

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Employee:
    def __init__(self, top=None):
        top.geometry("1300x600")
        # top.resizable(0, 0)
        top.title("EmprendeSoft - Administración - Empleados")
        top.state("zoomed")

        self.message = Label(emp)
        self.message.place(relx=0.035, rely=0.055, width=136, height=30)
        self.message.configure(font="-family {Poppins} -size 10")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(emp)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(emp)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        self.button1 = ttk.Button(
            top, text="Buscar ID", style="info-outline.TButton", command=self.search_emp
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Empleado",
            style="info-outline.TButton",
            command=self.add_emp,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Empleado",
            style="success-outline.TButton",
            command=self.update_emp,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Empleado",
            style="danger-outline.TButton",
            command=self.delete_emp,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(emp, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(emp, orient=VERTICAL)
        self.tree = ttk.Treeview(emp)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=450)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=750, height=22)

        self.tree.configure(
            columns=(
                "Employee ID",
                "Employee Name",
                "Contact No.",
                "Address",
                "Aadhar No.",
                "Password",
                "Designation",
            )
        )

        self.tree.heading("Employee ID", text="ID Empleado", anchor=W)
        self.tree.heading("Employee Name", text="Nombre", anchor=W)
        self.tree.heading("Contact No.", text="Teléfono de contacto", anchor=W)
        self.tree.heading("Address", text="Ubicación", anchor=W)
        self.tree.heading("Aadhar No.", text="Código", anchor=W)
        self.tree.heading("Password", text="Contraseña", anchor=W)
        self.tree.heading("Designation", text="Rol", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=198)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)
        self.tree.column("#7", stretch=NO, minwidth=0, width=80)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM employee")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_emp(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        to_search = self.entry1.get()
        for search in val:
            if search == to_search:
                self.tree.selection_set(val[val.index(search) - 1])
                self.tree.focus(val[val.index(search) - 1])
                messagebox.showinfo(
                    "Mensaje",
                    "ID Empleado: {} Encontrado".format(self.entry1.get()),
                    parent=emp,
                )
                break
        else:
            messagebox.showerror(
                "Error",
                "ID Empleado: {} no encontrado.".format(self.entry1.get()),
                parent=emp,
            )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_emp(self):
        val = []
        to_delete = []

        if len(self.sel) != 0:
            sure = messagebox.askyesno(
                "Confirmación",
                "¿Está seguro de eliminar empleado?",
                parent=emp,
            )
            if sure == True:
                for i in self.sel:
                    for j in self.tree.item(i)["values"]:
                        val.append(j)

                for j in range(len(val)):
                    if j % 7 == 0:
                        to_delete.append(val[j])

                for k in to_delete:
                    if k == "EMP0000":
                        flag = 0
                        break
                    else:
                        delete = "DELETE FROM employee WHERE emp_id = ?"
                        cur.execute(delete, [k])
                        db.commit()

                messagebox.showinfo(
                    "Mensaje", "Empleado eliminado de la base de datos.", parent=emp
                )
                self.sel.clear()
                self.tree.delete(*self.tree.get_children())
                self.DisplayData()

        else:
            messagebox.showerror(
                "Error", "Por favor seleccionar columna en la tabla", parent=emp
            )

    def update_emp(self):

        if len(self.sel) == 1:
            global e_update
            e_update = Toplevel()
            page8 = Update_Employee(e_update)
            page8.time()
            e_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global vall
            vall = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    vall.append(j)

            page8.entry1.insert(0, vall[1])
            page8.entry2.insert(0, vall[2])
            page8.entry3.insert(0, vall[4])
            page8.entry4.insert(0, vall[6])
            page8.entry5.insert(0, vall[3])
            # page8.entry6.insert(0, vall[5])
            e_update.mainloop()
        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccionar un empleado en la tabla"
            )
        else:
            messagebox.showerror("Error", "Can only update one employee at a time.")

    def add_emp(self):
        global e_add
        e_add = Toplevel()
        page6 = add_employee(e_add)
        page6.time()
        e_add.protocol("WM_DELETE_WINDOW", self.ex)
        e_add.mainloop()

    def ex(self):
        e_add.destroy()
        self.tree.delete(*self.tree.get_children())
        self.DisplayData()

    def ex2(self):
        e_update.destroy()
        self.tree.delete(*self.tree.get_children())
        self.DisplayData()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Salir", "¿Estás seguro de salir?", parent=emp)
        if sure == True:
            emp.destroy()
            adm.deiconify()

    def Logout(self):
        sure = messagebox.askyesno(
            "Cerrar Sesión", "¿Estás seguro de que quieres salir?"
        )
        if sure == True:
            emp.destroy()
            root.deiconify()

            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

class add_employee:
    # vista
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Empleado")

        self.r1 = e_add.register(self.testint)
        self.r2 = e_add.register(self.testchar)

        self.clock = ttk.Label(e_add)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")
        # self.clock.configure(background="#ffffff")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del Empleado", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Nombre del Empleado", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(e_add)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        # self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            top, text="N. Teléfono del Empleado", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(e_add)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        # self.entry2.configure(relief="flat")
        self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            top, text="Código", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(e_add)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        # self.entry3.configure(relief="flat")

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Rol o designación del empleado",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(e_add)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        # self.entry4.configure(relief="flat")
        self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(e_add)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para Contraseña
        self.label_password = Label(
            top, text="Contraseña", font=("Times", 15), foreground="#fff"
        )
        self.label_password.place(relx=0.527, rely=0.481)

        self.entry6 = Entry(e_add)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        # self.entry6.configure(relief="flat")
        self.entry6.configure(show="*")

        self.button_agregar = ttk.Button(
            e_add, text="Agregar", style="info.TButton", command=self.add
        )
        self.button_agregar.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button_agregar.configure(cursor="hand2")

        self.button_limpiar = ttk.Button(
            e_add, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button_limpiar.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button_limpiar.configure(cursor="hand2")

        self.button_code_generate = ttk.Button(
            e_add,
            text="Generar Código",
            style="success-outline.TButton",
            command=self.generar_y_colocar_codigo,
        )
        self.button_code_generate.place(relx=0.655, rely=0.836, width=120, height=35)
        self.button_code_generate.configure(cursor="hand2")

    # fin vista

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    # controlador
    def add(self):
        ename = self.entry1.get()
        econtact = self.entry2.get()
        eaddhar = self.entry3.get()
        edes = self.entry4.get()
        eadd = self.entry5.get()
        epass = self.entry6.get()

        # Verificar si se ingresó el nombre del empleado
        if ename.strip():
            # Verificar si se ingresó el rol del empleado
            if edes:
                # Verificar si se ingresó la dirección del empleado
                if eadd:
                    # Verificar si se ingresó la contraseña
                    if epass:
                        # Generar un hash de la contraseña
                        hashed_password = bcrypt.hash(epass)

                        # Generar un ID de empleado aleatorio
                        emp_id = random_emp_id(7)

                        # Consulta para insertar el empleado en la base de datos con la contraseña hasheada
                        insert = "INSERT INTO employee(emp_id, name, contact_num, address, aadhar_num, password, designation) VALUES(?,?,?,?,?,?,?)"
                        cur.execute(
                            insert,
                            [
                                emp_id,
                                ename,
                                econtact,
                                eadd,
                                eaddhar,
                                hashed_password,
                                edes,
                            ],
                        )
                        db.commit()

                        # Mostrar mensaje de éxito
                        messagebox.showinfo(
                            "Información",
                            "ID del empleado: {} ingresado correctamente".format(
                                emp_id
                            ),
                            parent=e_add,
                        )

                        # Limpiar los campos después de agregar el empleado
                        self.clearr()
                    else:
                        messagebox.showerror(
                            "Error", "Por favor ingrese la contraseña", parent=e_add
                        )
                else:
                    messagebox.showerror(
                        "Error",
                        "Por favor ingrese la dirección del empleado",
                        parent=e_add,
                    )
            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese el rol del empleado", parent=e_add
                )
        else:
            messagebox.showerror("Error", "Por favor ingresar nombre", parent=e_add)

    # fin controlador

    # Funciones para cifrar y verificar contraseñas con bcrypt
    def hash_password(password):
        hashed_password = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        return hashed_password

    def generar_y_colocar_codigo(self):
        longitud = 10
        # caracteres = string.ascii_letters + string.digits
        caracteres = string.digits
        codigo = "".join(random.choice(caracteres) for _ in range(longitud))
        self.entry3.delete(0, "end")  # Limpiar la entrada actual
        self.entry3.insert(0, codigo)  # Colocar el código generado en la entrada

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

class Update_Employee:
    def __init__(self, top=None, password=""):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Empleado")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(e_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Empleado", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.r1 = e_update.register(self.testint)
        self.r2 = e_update.register(self.testchar)

        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Nombre del Empleado", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(e_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            top, text="N. Teléfono del Empleado", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(e_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            top, text="Código", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(e_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Rol o designación del empleado",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(e_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(e_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para Contraseña
        self.label_password = Label(
            top, text="Contraseña", font=("Times", 15), foreground="#fff"
        )
        self.label_password.place(relx=0.527, rely=0.481)

        self.entry6 = Entry(e_update)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        self.entry6.configure(relief="flat")
        self.entry6.configure(show="*")
        self.entry6.insert(0, password)

        self.button1 = ttk.Button(
            e_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            e_update, text="Limpiar", style="danger.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def update(self):
        ename = self.entry1.get()
        econtact = self.entry2.get()
        eaddhar = self.entry3.get()
        edes = self.entry4.get()
        eadd = self.entry5.get()
        epass = self.entry6.get()

        if ename.strip() and edes and eadd and epass:
            # Genera el hash de la contraseña
            hashed_password = bcrypt.hash(epass.encode("utf-8"))

            emp_id = vall[0]
            update_query = "UPDATE employee SET name = ?, contact_num = ?, address = ?, aadhar_num = ?, password = ?, designation = ? WHERE emp_id = ?"
            cur.execute(
                update_query,
                [ename, econtact, eadd, eaddhar, hashed_password, edes, emp_id],
            )
            db.commit()
            messagebox.showinfo(
                "Success!!",
                "Employee ID: {} successfully updated in database.".format(emp_id),
                parent=e_update,
            )
            vall.clear()
            page5.tree.delete(*page5.tree.get_children())
            page5.DisplayData()
            Employee.sel.clear()
            e_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=e_add)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Invoice:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.state("zoomed")
        top.resizable(1, 1)
        top.title("EmprendeSoft - Facturas")

        # self.label1 = Label(invoice)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./invoices.png")
        # self.label1.configure(image=self.img)

        self.message = Label(invoice)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 10")
        self.message.configure(text=username)

        self.clock = Label(invoice)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(invoice)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        self.button1 = ttk.Button(
            top, text="Buscar ID", style="info-outline.TButton", command=self.search_inv
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Eliminar factura",
            style="danger.TButton",
            command=self.delete_invoice,
        )
        self.button3.place(relx=0.052, rely=0.432, width=306, height=28)
        self.button3.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(invoice, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(invoice, orient=VERTICAL)
        self.tree = ttk.Treeview(invoice)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=550)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<Double-1>", self.double_tap)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=548)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "Bill Number",
                "Date",
                "Customer Name",
                "Customer Phone No.",
            )
        )

        self.tree.heading("Bill Number", text="Bill Number", anchor=W)
        self.tree.heading("Date", text="Date", anchor=W)
        self.tree.heading("Customer Name", text="Customer Name", anchor=W)
        self.tree.heading("Customer Phone No.", text="Customer Phone No.", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=219)
        self.tree.column("#2", stretch=NO, minwidth=0, width=219)
        self.tree.column("#3", stretch=NO, minwidth=0, width=219)
        self.tree.column("#4", stretch=NO, minwidth=0, width=219)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM bill")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def double_tap(self, Event):
        item = self.tree.identify("item", Event.x, Event.y)
        global bill_num
        bill_num = self.tree.item(item)["values"][0]

        global bill
        bill = Toplevel()
        pg = open_bill(bill)
        # bill.protocol("WM_DELETE_WINDOW", exitt)
        bill.mainloop()

    def delete_invoice(self):
        val = []
        to_delete = []

        if len(self.sel) != 0:
            sure = messagebox.askyesno(
                "Confirmación",
                "¿Está seguro de eliminar la selección?",
                parent=invoice,
            )
            if sure == True:
                for i in self.sel:
                    for j in self.tree.item(i)["values"]:
                        val.append(j)

                for j in range(len(val)):
                    if j % 5 == 0:
                        to_delete.append(val[j])

                for k in to_delete:
                    delete = "DELETE FROM bill WHERE bill_no = ?"
                    cur.execute(delete, [k])
                    db.commit()

                messagebox.showinfo(
                    "Correcto",
                    "Facturas se eliminó de la base de datos",
                    parent=invoice,
                )
                self.sel.clear()
                self.tree.delete(*self.tree.get_children())

                self.DisplayData()
        else:
            messagebox.showerror("Error", "Por favor ", parent=invoice)

    def search_inv(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        to_search = self.entry1.get()
        for search in val:
            if search == to_search:
                self.tree.selection_set(val[val.index(search) - 1])
                self.tree.focus(val[val.index(search) - 1])
                messagebox.showinfo(
                    "Información",
                    "Número: {} Encontrado.".format(self.entry1.get()),
                    parent=invoice,
                )
                break
        else:
            messagebox.showerror(
                "Error",
                "Número: {} no encontrado".format(self.entry1.get()),
                parent=invoice,
            )

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Desea salir?")
        if sure == True:
            invoice.destroy()
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Salir", "¿Usted quiere salir?", parent=invoice)
        if sure == True:
            invoice.destroy()
            adm.deiconify()

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM bill"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/facturas_de_ventas.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()


class open_bill:
    def __init__(self, top=None):

        top.geometry("800x500")
        top.resizable(0, 0)
        top.title("Factura")

        self.name_message = Text(bill)
        self.name_message.place(relx=0.178, rely=0.205, width=176, height=30)
        self.name_message.configure(font="-family {Podkova} -size 10")
        self.name_message.configure(borderwidth=0)
        self.name_message.configure(background="#ffffff")

        self.num_message = Text(bill)
        self.num_message.place(relx=0.854, rely=0.205, width=90, height=30)
        self.num_message.configure(font="-family {Podkova} -size 10")
        self.num_message.configure(borderwidth=0)
        self.num_message.configure(background="#ffffff")

        self.bill_message = Text(bill)
        self.bill_message.place(relx=0.150, rely=0.243, width=176, height=26)
        self.bill_message.configure(font="-family {Podkova} -size 10")
        self.bill_message.configure(borderwidth=0)
        self.bill_message.configure(background="#ffffff")

        self.bill_date_message = Text(bill)
        self.bill_date_message.place(relx=0.780, rely=0.243, width=90, height=26)
        self.bill_date_message.configure(font="-family {Podkova} -size 10")
        self.bill_date_message.configure(borderwidth=0)
        self.bill_date_message.configure(background="#ffffff")

        self.Scrolledtext1 = tkst.ScrolledText(top)
        self.Scrolledtext1.place(relx=0.044, rely=0.41, width=695, height=284)
        self.Scrolledtext1.configure(borderwidth=0)
        self.Scrolledtext1.configure(font="-family {Podkova} -size 8")
        self.Scrolledtext1.configure(state="disabled")

        find_bill = "SELECT * FROM bill WHERE bill_no = ?"
        cur.execute(find_bill, [bill_num])
        results = cur.fetchall()
        if results:
            self.name_message.insert(END, results[0][2])
            self.name_message.configure(state="disabled")

            self.num_message.insert(END, results[0][3])
            self.num_message.configure(state="disabled")

            self.bill_message.insert(END, results[0][0])
            self.bill_message.configure(state="disabled")

            self.bill_date_message.insert(END, results[0][1])
            self.bill_date_message.configure(state="disabled")

            self.Scrolledtext1.configure(state="normal")
            self.Scrolledtext1.insert(END, results[0][4])
            self.Scrolledtext1.configure(state="disabled")

class Client:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.state("zoomed")
        top.resizable(1, 1)
        top.title("Clientes")

        # self.label1 = Label(invoice)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./invoices.png")
        # self.label1.configure(image=self.img)

        self.message = Label(client)
        self.message.place(relx=0.046, rely=0.055, width=136, height=30)
        self.message.configure(font="-family {Poppins} -size 10")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(client)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(client)
        self.entry1.place(relx=0.040, rely=0.286, width=240, height=28)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        self.button1 = ttk.Button(
            top, text="Buscar", style="info.TButton", command=self.search_client
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")


        self.button_add = ttk.Button(
            top,
            text="Agregar Cliente",
            style="info-outline.TButton",
            command=self.add_client,
        )
        self.button_add.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button_add.configure(cursor="hand2")

        self.button_update = ttk.Button(
            top,
            text="Actualizar Cliente",
            style="success-outline.TButton",
            command=self.update_client,
        )
        self.button_update.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button_update.configure(cursor="hand2")

        self.button_delete = ttk.Button(
            top,
            text="Eliminar Cliente",
            style="danger-outline.TButton",
            command=self.delete_client,
        )
        self.button_delete.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button_delete.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(client, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(client, orient=VERTICAL)
        self.tree = ttk.Treeview(client)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Name",
                "Last",
                "Phone",
                "Address",
                "Email",
                "Type"
            )
        )

        self.tree.heading("ID", text="ID ", anchor=W)
        self.tree.heading("Name", text="Nombre", anchor=W)
        self.tree.heading("Last", text="Apellidos", anchor=W)
        self.tree.heading("Phone", text="Tel. Contacto", anchor=W)
        self.tree.heading("Address", text="Dirección", anchor=W)
        self.tree.heading("Email", text="Correo", anchor=W)
        self.tree.heading("Type", text="Tipo", anchor=W)


        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)

        self.DisplayData()

    def add_client(self):
        global e_add
        global page6
        e_add = Toplevel()
        page6 = add_client(e_add)
        page6.time()
        e_add.mainloop()

    sel = []

    def on_tree_select(self, event):
        # Limpiar la lista de selección
        self.sel.clear()
        # Obtener las filas seleccionadas en la tabla
        selected_items = self.tree.selection()
        for item in selected_items:
            # Obtener los valores de cada fila seleccionada
            values = self.tree.item(item, option="values")
            # Agregar los valores a la lista de selección
            self.sel.append(values)

    def update_client(self):

        if len(self.sel) == 1:
            global Client_update
            Client_update = Toplevel()
            page_client = Update_Client(Client_update)
            
            # Si se selecciona una sola fila, obtener los datos de esa fila
            selected_data = self.sel[0]  # Obtener la primera fila seleccionada
            # Mostrar los datos en el formulario de actualización
            page_client.entry_id.insert(0, selected_data[0])
            page_client.entry_nombre.insert(0, selected_data[1])
            page_client.entry2.insert(0, selected_data[2])
            page_client.entry3.insert(0, selected_data[3])
            page_client.entry4.insert(0, selected_data[4])
            page_client.entry5.insert(0, selected_data[5])
            page_client.entry6.insert(0, selected_data[6])
            # Mostrar la ventana de actualización
            Client_update.mainloop()

      

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=client
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=client
            )

    def DisplayData(self):
        cur.execute("SELECT * FROM client")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))
   
    def delete_client(self):
        selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione un cliente para eliminar.")
            return

        # Obtener el ID y otros datos del cliente seleccionado
        client_data = self.tree.item(selected_item)['values']
        client_id = client_data[0]  # ID del cliente

        # Mostrar confirmación antes de eliminar
        confirm = messagebox.askyesno(
        "Confirmar Eliminación",
        f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
        )
        if confirm:
            try:
                cur.execute("DELETE FROM client WHERE id = ?", (client_id,))
                db.commit()
                messagebox.showinfo("Éxito", "Cliente eliminado correctamente.")
                page8.tree.delete(*page8.tree.get_children())
                self.DisplayData()  # Actualizar la tabla después de eliminar
            except sqlite3.Error as e:
                messagebox.showerror("Error", f"Error al eliminar el cliente: {e}")
        

    def search_client(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "ID Cliente invalido", parent=client)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID Cliente: {} encontrado.".format(self.entry1.get()),
                        parent=client,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Cliente: {} no encontrado".format(self.entry1.get()),
                    parent=client,
                )

    sel = []

    def Logout(self):
        sure = messagebox.askyesno("Logout", "Are you sure you want to logout?")
        if sure == True:
            invoice.destroy()
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno(
            "Panel de control", "¿Quiere regresar al panel de control?", parent=client
        )
        if sure == True:
            client.destroy()
            adm.deiconify()

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM client"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/clientes.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_client:
    # vista
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Cliente")

        self.r1 = e_add.register(self.testint)
        self.r2 = e_add.register(self.testchar)

        self.clock = ttk.Label(e_add)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")
        # self.clock.configure(background="#ffffff")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del Cliente", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(e_add)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        # self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            top, text="Apellidos", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(e_add)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        # self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            top, text="Tel. Contacto", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(e_add)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        # self.entry3.configure(relief="flat")

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Dirección",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(e_add)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        # self.entry4.configure(relief="flat")
        self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Correo", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(e_add)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para Contraseña
        self.label_password = Label(
            top, text="Tipo de Cliente", font=("Times", 15), foreground="#fff"
        )
        self.label_password.place(relx=0.527, rely=0.481)

        self.entry6 = Entry(e_add)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        # self.entry6.configure(relief="flat")
        #self.entry6.configure(show="*")

        self.button_agregar = ttk.Button(
            e_add, text="Agregar", style="info.TButton", command=self.add
        )
        self.button_agregar.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button_agregar.configure(cursor="hand2")

        self.button_limpiar = ttk.Button(
            e_add, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button_limpiar.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button_limpiar.configure(cursor="hand2")

    # fin vista

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    # controlador
    def add(self):
        name = self.entry1.get()
        last_name= self.entry2.get()
        contact_num = self.entry3.get()
        address = self.entry4.get()
        email = self.entry5.get()
        type_client = self.entry6.get()

        # Verificar si se ingresó el nombre del empleado
        if name.strip():
            # Verificar si se ingresó el rol del empleado
            if last_name.strip():
                # Verificar si se ingresó la dirección del empleado
                if contact_num.strip():
                    # Verificar si se ingresó la contraseña
                    if address.strip():

                        # Consulta para insertar el empleado en la base de datos con la contraseña hasheada
                        insert = "INSERT INTO client(name, last_name, contact_num, address, email, type_client) VALUES(?,?,?,?,?,?)"
                        cur.execute(
                            insert,
                            [
                                name,
                                last_name,
                                contact_num,
                                address,
                                email,
                                type_client,
                            ]
                        )
                        db.commit()
                        messagebox.showinfo(
                        "Información",
                        "Producto ingresado correctamente",
                        parent=e_add,
                        )
                        e_add.destroy()
                        page8.tree.delete(*page8.tree.get_children())
                        page8.DisplayData()
                        e_add.destroy()

                        # Limpiar los campos después de agregar el empleado
                        self.clearr()
                    else:
                        messagebox.showerror(
                            "Error", "Por favor ingrese la contraseña", parent=e_add
                        )
                else:
                    messagebox.showerror(
                        "Error",
                        "Por favor ingrese la dirección del empleado",
                        parent=e_add,
                    )
            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese el rol del empleado", parent=e_add
                )
        else:
            messagebox.showerror("Error", "Por favor ingresar nombre", parent=e_add)

    # fin controlador

    # Funciones para cifrar y verificar contraseñas con bcrypt
    def hash_password(password):
        hashed_password = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        return hashed_password

    def generar_y_colocar_codigo(self):
        longitud = 10
        # caracteres = string.ascii_letters + string.digits
        caracteres = string.digits
        codigo = "".join(random.choice(caracteres) for _ in range(longitud))
        self.entry3.delete(0, "end")  # Limpiar la entrada actual
        self.entry3.insert(0, codigo)  # Colocar el código generado en la entrada

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

    def DisplayData(self):
        cur.execute("SELECT * FROM client")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

class Update_Client:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Cliente")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Client_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Cliente", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Client_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)


        self.entry_nombre = Entry(Client_update)
        self.entry_nombre.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry_nombre.configure(font="-family {Poppins} -size 12")
        #self.entry_nombre.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Client_update, text="Apellidos", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Client_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Client_update, text="Teléfono", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Client_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Dirección",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Client_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Correo", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(Client_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Tipo", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.480)

        self.entry6 = Entry(Client_update)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        self.entry6.configure(relief="flat")

        self.button1 = ttk.Button(
            Client_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

        self.button_limpiar = ttk.Button(
            Client_update, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button_limpiar.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button_limpiar.configure(cursor="hand2")

    def update(self):
        # Obtener los valores actualizados del formulario
        id_cliente = self.entry_id.get()
        nombre = self.entry_nombre.get()
        apellidos = self.entry2.get()
        telefono = self.entry3.get()
        direccion = self.entry4.get()
        correo = self.entry5.get()
        tipo = self.entry6.get()
        if nombre.strip():
            update_query = "UPDATE client SET name = ?, last_name = ?, contact_num = ?, address = ?, email = ?, type_client = ? WHERE id = ?"
            cur.execute(
                update_query,
                [nombre, apellidos, telefono, direccion, correo, tipo, id_cliente],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Cliente ID: {} Actualizado Correctamente".format(id_cliente),
                parent=Client_update,
            )
            Client_update.destroy()
            page8.tree.delete(*page8.tree.get_children())
            page8.DisplayData()
        else:
            messagebox.showerror("Error", "Capos vacíos", parent=client)
       
 

        # Cerrar la ventana de actualización después de actualizar los datos
        


    def clearr(self):
        self.entry_nombre.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

class Provider:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Proveedor")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(provi)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(provi)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_product,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Proveedor",
            style="info-outline.TButton",
            command=self.add_providerr,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Proveedor",
            style="success-outline.TButton",
            command=self.update_provider,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Proveedor",
            style="danger-outline.TButton",
            command=self.delete_provider,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(provi, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(provi, orient=VERTICAL)
        self.tree = ttk.Treeview(provi)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Rut",
                "Name",
                "Email",
                "Phone",
                "Address"
            )
        )

        self.tree.heading("ID", text="ID Proveedor", anchor=W)
        self.tree.heading("Rut", text="Rut", anchor=W)
        self.tree.heading("Name", text="Nombre del proveedor", anchor=W)
        self.tree.heading("Email", text="Correo", anchor=W)
        self.tree.heading("Phone", text="Teléfono", anchor=W)
        self.tree.heading("Address", text="Localidad", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM provider")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_product(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id del producto invalido.", parent=inv)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID Producto: {} encontradp.".format(self.entry1.get()),
                        parent=inv,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} not found.".format(self.entry1.get()),
                    parent=inv,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_provider(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione un cliente para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            client_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM provider WHERE id = ?", (client_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Proveedor eliminado correctamente.")
                    page3.tree.delete(*page3.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar el proveedor: {e}")

    def update_provider(self):
        if len(self.sel) == 1:
            global Provider_update
            Provider_update = Toplevel()
            page9 = Update_Provider(Provider_update)
            page9.time()
            Provider_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page9.entry_id.insert(0, valll[0])
            page9.entry1.insert(0, valll[1])
            page9.entry2.insert(0, valll[2])
            page9.entry3.insert(0, valll[3])
            page9.entry4.insert(0, valll[4])
            page9.entry5.insert(0, valll[5])
            Provider_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=inv
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=inv
            )

    def add_providerr(self):
        global p_add
        global page4
        p_add = Toplevel()
        page4 = add_provider(p_add)
        page4.time()
        p_add.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=provi)
        if sure == True:
            provi.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM provider"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/provedores.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_provider:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Proveedor")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(p_add)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del Proveedor", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Rut", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(p_add)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(p_add)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        self.r2 = p_add.register(self.testint)

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Correo", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(p_add)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        # Etiqueta para mrp
        self.label_phone = Label(
            top, text="Teléfono", font=("Times", 15), foreground="#fff"
        )

        self.label_phone.place(relx=0.132, rely=0.590)

        self.entry4 = Entry(p_add)
        self.entry4.place(relx=0.132, rely=0.646, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Dirección o Localidad", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.370)

        self.entry6 = Entry(p_add)
        self.entry6.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

        self.button_calcular_total = ttk.Button(
            top,
            text="Calcular total",
            style="danger.TButton",
            command=self.calculartotal,
        )
        self.button_calcular_total.place(relx=0.640, rely=0.836, width=100, height=34)
        self.button_calcular_total.configure(cursor="hand2")

    def add(self):
        email = self.entry3.get()
        name = self.entry2.get()
        phone = self.entry4.get()
        rut = self.entry1.get()
        address = self.entry6.get()
        if rut.strip():
            if name.strip():
                if address.strip():
                    
                    insert = "INSERT INTO provider(rut, name, email, phone, address) VALUES(?,?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            rut,
                            name,
                            email,
                            phone,
                            address,
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Producto ingresado correctamente",
                        parent=p_add,
                    )
                    p_add.destroy()
                    page3.tree.delete(*page3.tree.get_children())
                    page3.DisplayData()
                    p_add.destroy()

                else:
                    messagebox.showerror(
                        "Error", "Por favor ingrese categoría", parent=p_add
                    )
            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=p_add
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=p_add)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Provider:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Proveedor")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Provider_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Proveedor", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Provider_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Rut", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Provider_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Provider_update, text="Nombre del proveedor", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Provider_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Provider_update, text="Correo", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Provider_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Teléfono",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Provider_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(Provider_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        self.button1 = ttk.Button(
            Provider_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_provider = self.entry_id.get()
        rut = self.entry1.get()
        name = self.entry2.get()
        email = self.entry3.get()
        phone = self.entry4.get()
        address = self.entry5.get()

        if rut.strip() and name and email and phone:
            update_query = "UPDATE provider SET rut = ?, name = ?, email = ?, phone = ?, address = ? WHERE id = ?"
            cur.execute(
                update_query,
                [rut, name, email, phone, address, id_provider],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Proveedor ID: {} Actualizado correctamenete".format(id_provider),
                parent=Provider_update,
            )
            page3.tree.delete(*page3.tree.get_children())
            page3.DisplayData()
            Provider.sel.clear()
            Provider_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Provider_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Transaction:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Transacciones")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(trans)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(trans)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Proveedor",
            style="info-outline.TButton",
            command=self.add_trans,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Proveedor",
            style="success-outline.TButton",
            command=self.update_trans,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Proveedor",
            style="danger-outline.TButton",
            command=self.delete_provider,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(trans, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(trans, orient=VERTICAL)
        self.tree = ttk.Treeview(trans)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "transaction_id",
                "date_transaction",
                "amount",
                "transaction_type",
                "entity",
                "payment_type",
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("transaction_id", text="ID transacción", anchor=W)
        self.tree.heading("date_transaction", text="Fecha", anchor=W)
        self.tree.heading("amount", text="Monto", anchor=W)
        self.tree.heading("transaction_type", text="Tipo de transaccion", anchor=W)
        self.tree.heading("entity", text="Entidad", anchor=W)
        self.tree.heading("payment_type", text="Tipo de pago", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)
        self.tree.column("#7", stretch=NO, minwidth=0, width=80)


        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM transactions")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=trans)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Transacción: {} encontrado.".format(self.entry1.get()),
                        parent=trans,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} not found.".format(self.entry1.get()),
                    parent=page_transaction,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_provider(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione un cliente para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            client_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM transactions WHERE id = ?", (client_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Proveedor eliminado correctamente.")
                    page_transaction.tree.delete(*page_transaction.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar el proveedor: {e}")

    def update_trans(self):
        if len(self.sel) == 1:
            global Transaction_update
            Transaction_update = Toplevel()
            page9 = Update_Transaction(Transaction_update)
            page9.time()
            Transaction_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page9.entry_id.insert(0, valll[0])
            page9.entry1.insert(0, valll[1])
            page9.entry2.insert(0, valll[2])
            page9.entry3.insert(0, valll[3])
            page9.entry4.insert(0, valll[4])
            page9.entry5.insert(0, valll[5])
            page9.entry6.insert(0, valll[6])
            Transaction_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=inv
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=inv
            )

    def add_trans(self):
        global page_add_trans
        global page4
        page_add_trans = Toplevel()
        page4 = add_transaction(page_add_trans)
        page4.time()
        page_add_trans.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=trans)
        if sure == True:
            trans.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM transactions"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/transacciones.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_transaction:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Transacciones")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_trans)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro de la transacción", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Id transacción", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(page_add_trans)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Fecha", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(page_add_trans)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Monto", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(page_add_trans)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        # Etiqueta para mrp
        self.label_phone = Label(
            top, text="Tipo de transacción", font=("Times", 15), foreground="#fff"
        )

        self.label_phone.place(relx=0.132, rely=0.590)

        self.entry4 = Entry(page_add_trans)
        self.entry4.place(relx=0.132, rely=0.646, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_address = Label(
            page_add_trans, text="Entidad", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.255)

        self.entry5 = Entry(page_add_trans)
        self.entry5.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Tipo de pago", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.370)

        self.entry6 = Entry(page_add_trans)
        self.entry6.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

        self.button_calcular_total = ttk.Button(
            top,
            text="Calcular total",
            style="danger.TButton",
            command=self.calculartotal,
        )
        self.button_calcular_total.place(relx=0.640, rely=0.836, width=100, height=34)
        self.button_calcular_total.configure(cursor="hand2")

    def add(self):
        id_trans = self.entry1.get()
        date = self.entry2.get()
        mount = self.entry3.get()
        type_trans = self.entry4.get()
        entity = self.entry5.get()
        type_payment = self.entry6.get()

        if id_trans.strip():
            if date.strip():
                if mount.strip():
                    insert = "INSERT INTO transactions (transaction_id, date_transaction, amount, transaction_type, entity, payment_type) VALUES(?,?,?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            id_trans,
                            date,
                            mount,
                            type_trans,
                            entity,
                            type_payment,
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Producto ingresado correctamente",
                        parent=page_add_trans,
                    )
                    page_add_trans.destroy()
                    page_transaction.tree.delete(*page_transaction.tree.get_children())
                    page_transaction.DisplayData()
                    page_add_trans.destroy()

                else:
                    messagebox.showerror(
                        "Error", "Por favor ingrese categoría", parent=page_add_trans
                    )
            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_trans
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_trans)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Transaction:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Transacción")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Transaction_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Transacción", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Transaction_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Rut", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Transaction_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Transaction_update, text="Nombre del proveedor", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Transaction_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Transaction_update, text="Correo", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Transaction_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Teléfono",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Transaction_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(Transaction_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Tipo de pago", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.480)

        self.entry6 = Entry(Transaction_update)
        self.entry6.place(relx=0.527, rely=0.580, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")
        self.entry6.configure(relief="flat")


        self.button1 = ttk.Button(
            Transaction_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_transaction = self.entry_id.get()
        code = self.entry1.get()
        date = self.entry2.get()
        mount = self.entry3.get()
        type_trans = self.entry4.get()
        entity = self.entry5.get()
        type_pay = self.entry6.get()

        if code.strip() and date and mount and type_trans:
            update_query = "UPDATE transactions SET transaction_id = ?, date_transaction = ?, amount = ?, transaction_type = ?, entity = ?, payment_type = ? WHERE id = ?"
            cur.execute(
                update_query,
                [code, date, mount, type_trans, entity, type_pay, id_transaction],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Proveedor ID: {} Actualizado correctamenete".format(id_transaction),
                parent=Transaction_update,
            )
            page_transaction.tree.delete(*page_transaction.tree.get_children())
            page_transaction.DisplayData()
            Transaction.sel.clear()
            Transaction_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Transaction_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Project:

    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Proyectos")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(pro)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(pro)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Proveedor",
            style="info-outline.TButton",
            command=self.add_trans,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Proveedor",
            style="success-outline.TButton",
            command=self.update_trans,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Proveedor",
            style="danger-outline.TButton",
            command=self.delete_project,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(pro, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(pro, orient=VERTICAL)
        self.tree = ttk.Treeview(pro)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Name",
                "Description",
                "Budget",
                "status",
                "type_project",
                "payment_type",
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("Name", text="Nombre", anchor=W)
        self.tree.heading("Description", text="Descripción", anchor=W)
        self.tree.heading("Budget", text="Precio estimado", anchor=W)
        self.tree.heading("status", text="Estado", anchor=W)
        self.tree.heading("type_project", text="Tipo", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)


        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM project")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=trans)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Transacción: {} encontrado.".format(self.entry1.get()),
                        parent=trans,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} not found.".format(self.entry1.get()),
                    parent=page_transaction,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_project(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione un cliente para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            project_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM project WHERE id = ?", (project_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Proyecto eliminado correctamente.")
                    page_project.tree.delete(*page_project.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar el proveedor: {e}")

    def update_trans(self):
        if len(self.sel) == 1:
            global Project_update
            Project_update = Toplevel()
            page9 = Update_Project(Project_update)
            page9.time()
            Project_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page9.entry_id.insert(0, valll[0])
            page9.entry1.insert(0, valll[1])
            page9.entry2.insert(0, valll[2])
            page9.entry3.insert(0, valll[3])
            page9.entry4.insert(0, valll[4])
            page9.entry5.insert(0, valll[5])
            Transaction_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=inv
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=inv
            )

    def add_trans(self):
        global page_add_project
        global pageproject
        page_add_project = Toplevel()
        pageproject = add_project(page_add_project)
        pageproject.time()
        page_add_project.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=pro)
        if sure == True:
            pro.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM project"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/proyectos.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_project:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Proyecto")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_project)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del proyecto", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Nombre", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(page_add_project)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(page_add_project)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Estimación gastos", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(page_add_project)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        # Etiqueta para mrp
        self.label_phone = Label(
            top, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_phone.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(page_add_project)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Tipo de proyecto", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.370)

        self.entry5 = Entry(page_add_project)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

  

    def add(self):
        name = self.entry1.get()
        description = self.entry2.get()
        budget = self.entry3.get()
        status = self.entry4.get()
        type_project = self.entry5.get()
        if name.strip():
            if description.strip():
                    insert = "INSERT INTO project(name, description, budget, status, type_project) VALUES(?,?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            name,
                            description,
                            budget,
                            status,
                            type_project
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Proyecto ingresado correctamente",
                        parent=page_add_project,
                    )
                    page_add_project.destroy()
                    page_project.tree.delete(*page_project.tree.get_children())
                    page_project.DisplayData()
                    page_add_project.destroy()

            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_project
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_project)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Project:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Proyecto")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Project_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Transacción", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Project_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Rut", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Project_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Project_update, text="Nombre del proveedor", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Project_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Project_update, text="Correo", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Project_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Teléfono",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Project_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(Project_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")


        self.button1 = ttk.Button(
            Project_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_project = self.entry_id.get()
        name = self.entry1.get()
        description = self.entry2.get()
        budget = self.entry3.get()
        status = self.entry4.get()
        type_project = self.entry5.get()
        

        if name.strip() and description and budget and status:
            update_query = "UPDATE project SET name = ?, description = ?, budget = ?, status = ?, type_project = ? WHERE id = ?"
            cur.execute(
                update_query,
                [name, description, budget, status, type_project, id_project],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Proveedor ID: {} Actualizado correctamenete".format(id_project),
                parent=Project_update,
            )
            page_project.tree.delete(*page_project.tree.get_children())
            page_project.DisplayData()
            Transaction.sel.clear()
            Project_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Project_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Sent:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Envíos")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(se)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(se)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Proveedor",
            style="info-outline.TButton",
            command=self.add_sent,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Proveedor",
            style="success-outline.TButton",
            command=self.update_trans,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Proveedor",
            style="danger-outline.TButton",
            command=self.delete_project,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(se, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(se, orient=VERTICAL)
        self.tree = ttk.Treeview(se)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "address",
                "description",
                "price_total",
                "status",
                "product",
                "quantity_product",
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("address", text="Address", anchor=W)
        self.tree.heading("description", text="Descripción", anchor=W)
        self.tree.heading("price_total", text="Precio Total", anchor=W)
        self.tree.heading("status", text="Estado", anchor=W)
        self.tree.heading("product", text="Producto", anchor=W)
        self.tree.heading("quantity_product", text="Cantidad de productos", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)
        self.tree.column("#6", stretch=NO, minwidth=0, width=80)
        self.tree.column("#7", stretch=NO, minwidth=0, width=80)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM sent")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=trans)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Transacción: {} encontrado.".format(self.entry1.get()),
                        parent=trans,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} not found.".format(self.entry1.get()),
                    parent=page_transaction,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_project(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione un envío para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            sent_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM sent WHERE id = ?", (sent_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Proyecto eliminado correctamente.")
                    page_sent.tree.delete(*page_sent.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar el proveedor: {e}")

    def update_trans(self):
        if len(self.sel) == 1:
            global Sent_update
            Sent_update = Toplevel()
            page_sent = Update_Sent(Sent_update)
            page_sent.time()
            Sent_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page_sent.entry_id.insert(0, valll[0])
            page_sent.entry1.insert(0, valll[1])
            page_sent.entry2.insert(0, valll[2])
            page_sent.entry3.insert(0, valll[3])
            page_sent.entry4.insert(0, valll[4])
            page_sent.entry5.insert(0, valll[5])
            page_sent.entry6.insert(0, valll[6])
            Sent_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=se
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=se
            )

    def add_sent(self):
        global page_add_sent
        global pagesent
        page_add_sent = Toplevel()
        pagesent = add_sent(page_add_sent)
        pagesent.time()
        page_add_sent.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=se)
        if sure == True:
            se.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM sent"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/envios.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_sent:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Envíos")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_sent)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro del Envío", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Address", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(page_add_sent)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(page_add_sent)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Precio Total", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(page_add_sent)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        # Etiqueta para mrp
        self.label_phone = Label(
            top, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_phone.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(page_add_sent)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Producto", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.370)

        self.entry5 = Entry(page_add_sent)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Cantidad", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.481)

        self.entry6 = Entry(page_add_sent)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def add(self):
        address = self.entry1.get()
        description = self.entry2.get()
        price_total = self.entry3.get()
        status = self.entry4.get()
        product = self.entry5.get()
        quantity_product = self.entry6.get()

        if address.strip():
            if description.strip():
                    insert = "INSERT INTO sent(address, description, price_total, status, product, quantity_product) VALUES(?,?,?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            address,
                            description,
                            price_total,
                            status,
                            product,
                            quantity_product
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Proyecto ingresado correctamente",
                        parent=page_add_sent,
                    )
                    page_add_sent.destroy()
                    page_sent.tree.delete(*page_sent.tree.get_children())
                    page_sent.DisplayData()
                    page_add_sent.destroy()

            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_sent
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_sent)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Sent:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Envío")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Sent_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Envío", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Sent_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Sent_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Sent_update, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Sent_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Sent_update, text="Precio Total de estimación", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Sent_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Estado",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Sent_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        # Etiqueta para Dirección del empleado
        self.label_addresss = Label(
            top, text="Producto", font=("Times", 15), foreground="#fff"
        )
        self.label_addresss.place(relx=0.527, rely=0.375)

        self.entry5 = Entry(Sent_update)
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")

        # Etiqueta para sub category
        self.label_address = Label(
            top, text="Cantidad", font=("Times", 15), foreground="#fff"
        )

        self.label_address.place(relx=0.527, rely=0.481)

        self.entry6 = Entry(Sent_update)
        self.entry6.place(relx=0.527, rely=0.529, width=374, height=30)
        self.entry6.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            Sent_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_sent = self.entry_id.get()
        address = self.entry1.get()
        description = self.entry2.get()
        price_total = self.entry3.get()
        status = self.entry4.get()
        name_product = self.entry5.get()
        quantity_product = self.entry6.get()

        if address.strip() and description and price_total and status:
            update_query = "UPDATE sent SET address = ?, description = ?, price_total = ?, status = ?, product = ?, quantity_product = ? WHERE id = ?"
            cur.execute(
                update_query,
                [address, description, price_total, status, name_product,quantity_product, id_sent],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Envío ID: {} Actualizado correctamenete".format(id_sent),
                parent=Sent_update,
            )
            page_sent.tree.delete(*page_sent.tree.get_children())
            page_sent.DisplayData()
            Sent.sel.clear()
            Sent_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Sent_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Investment:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Inversiones")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(inves)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(inves)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Proveedor",
            style="info-outline.TButton",
            command=self.add_sent,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Proveedor",
            style="success-outline.TButton",
            command=self.update_inves,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Proveedor",
            style="danger-outline.TButton",
            command=self.delete_project,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(inves, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(inves, orient=VERTICAL)
        self.tree = ttk.Treeview(inves)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Type",
                "Date",
                "Yield",
                "Expiration",
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("Type", text="Tipo", anchor=W)
        self.tree.heading("Date", text="Fecha", anchor=W)
        self.tree.heading("Yield", text="Monto", anchor=W)
        self.tree.heading("Expiration", text="Fecha de expiración", anchor=W)


        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=120)
        self.tree.column("#5", stretch=NO, minwidth=0, width=80)

        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM investments")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=inves)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Transacción: {} encontrado.".format(self.entry1.get()),
                        parent=inves,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Producto: {} not found.".format(self.entry1.get()),
                    parent=inves,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_project(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione una inversión para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            inv_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM investments WHERE id = ?", (inv_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Inversión eliminada correctamente.")
                    page_investment.tree.delete(*page_investment.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar la inversión: {e}")

    def update_inves(self):
        if len(self.sel) == 1:
            global Investment_update
            Investment_update = Toplevel()
            page_investment = Update_Investment(Investment_update)
            page_investment.time()
            Investment_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page_investment.entry_id.insert(0, valll[0])
            page_investment.entry1.insert(0, valll[1])
            page_investment.entry2.insert(0, valll[2])
            page_investment.entry3.insert(0, valll[3])
            page_investment.entry4.insert(0, valll[4])
            Investment_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=inves
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=inves
            )

    def add_sent(self):
        global page_add_investment
        global pageinvestment
        page_add_investment = Toplevel()
        pageinvestment = add_investment(page_add_investment)
        pageinvestment.time()
        page_add_investment.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=inves)
        if sure == True:
            inves.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM investments"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/inversiones.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_investment:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Inversión")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_investment)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro de la Inversión", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Tipo", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(page_add_investment)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Fecha de la inversión", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(page_add_investment)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Monto Total", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(page_add_investment)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        # Etiqueta para mrp
        self.label_phone = Label(
            top, text="Fecha de expiración", font=("Times", 15), foreground="#fff"
        )

        self.label_phone.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(page_add_investment)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def add(self):
        type = self.entry1.get()
        date = self.entry2.get()
        mount_total = self.entry3.get()
        date_expiration = self.entry4.get()


        if type.strip():
            if date.strip():
                    insert = "INSERT INTO investments(type, date, yield, expiration_date) VALUES(?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            type,
                            date,
                            mount_total,
                            date_expiration,
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Proyecto ingresado correctamente",
                        parent=page_add_investment,
                    )
                    page_add_investment.destroy()
                    page_investment.tree.delete(*page_investment.tree.get_children())
                    page_investment.DisplayData()
                    page_add_investment.destroy()

            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_investment
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_investment)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Investment:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Inversión")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Investment_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Envío", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Investment_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Dirección", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Investment_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Investment_update, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Investment_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Investment_update, text="Precio Total de estimación", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Investment_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para Rol del empleado
        self.label_aadharno = Label(
            top,
            text="Estado",
            font=("Times", 15),
            foreground="#fff",
        )
        self.label_aadharno.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Investment_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry4.configure(validate="key", validatecommand=(self.r2, "%P"))

        self.button1 = ttk.Button(
            Investment_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_investments = self.entry_id.get()
        type = self.entry1.get()
        date = self.entry2.get()
        yield_inv = self.entry3.get()
        expiration_date = self.entry4.get()
 

        if type.strip() and date and yield_inv and expiration_date:
            update_query = "UPDATE investments SET type = ?, date = ?, yield = ?, expiration_date = ? WHERE id = ?"
            cur.execute(
                update_query,
                [type, date, yield_inv, expiration_date, id_investments],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Envío ID: {} Actualizado correctamenete".format(id_investments),
                parent=Investment_update,
            )
            page_investment.tree.delete(*page_investment.tree.get_children())
            page_investment.DisplayData()
            Investment.sel.clear()
            Investment_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Sent_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Task:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Tareas")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(taskk)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(taskk)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Tareas",
            style="info-outline.TButton",
            command=self.add_sent,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Tareas",
            style="success-outline.TButton",
            command=self.update_inves,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Tareas",
            style="danger-outline.TButton",
            command=self.delete_project,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(taskk, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(taskk, orient=VERTICAL)
        self.tree = ttk.Treeview(taskk)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Name",
                "Description",
                "Status",
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("Name", text="Nombre", anchor=W)
        self.tree.heading("Description", text="Descripción", anchor=W)
        self.tree.heading("Status", text="Estado", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM task")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=taskk)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Tarea: {} encontrado.".format(self.entry1.get()),
                        parent=inves,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Tarea: {} not found.".format(self.entry1.get()),
                    parent=inves,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_project(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione una inversión para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            inv_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM task WHERE id = ?", (inv_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Inversión eliminada correctamente.")
                    page_task.tree.delete(*page_task.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar la inversión: {e}")

    def update_inves(self):
        if len(self.sel) == 1:
            global Task_update
            Task_update = Toplevel()
            page_task = Update_Task(Task_update)
            page_task.time()
            Task_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page_task.entry_id.insert(0, valll[0])
            page_task.entry1.insert(0, valll[1])
            page_task.entry2.insert(0, valll[2])
            page_task.entry3.insert(0, valll[3])
            page_task.entry4.insert(0, valll[4])
            Task_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=taskk
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=taskk
            )

    def add_sent(self):
        global page_add_task
        global pagetask
        page_add_task = Toplevel()
        pagetask = add_task(page_add_task)
        pagetask.time()
        page_add_task.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=taskk)
        if sure == True:
            taskk.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=p_update)
        if sure == True:
            p_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM task"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/tareas.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_task:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Tareas")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_task)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro de la Tarea", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Nombre", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(page_add_task)
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = Entry(page_add_task)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = Entry(page_add_task)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def add(self):
        name = self.entry1.get()
        description = self.entry2.get()
        status = self.entry3.get()


        if name.strip():
            if description.strip():
                    insert = "INSERT INTO task(name, description, status) VALUES(?,?,?)"
                    cur.execute(
                        insert,
                        [
                            name,
                            description,
                            status,
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Tarea ingresada correctamente",
                        parent=page_add_task,
                    )
                    page_add_task.destroy()
                    page_task.tree.delete(*page_task.tree.get_children())
                    page_task.DisplayData()
                    page_add_task.destroy()

            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_task
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_task)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Task:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Tarea")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Task_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Actualizar Tarea", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Task_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            top, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Task_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Task_update, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Task_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Task_update, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Task_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        self.button1 = ttk.Button(
            Task_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_task = self.entry_id.get()
        name = self.entry1.get()
        description = self.entry2.get()
        status = self.entry3.get()
        
 

        if name.strip() and description and status:
            update_query = "UPDATE task SET name = ?, description = ?, status = ? WHERE id = ?"
            cur.execute(
                update_query,
                [name, description, status, id_task],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Envío ID: {} Actualizado correctamenete".format(id_task),
                parent=Task_update,
            )
            page_task.tree.delete(*page_task.tree.get_children())
            page_task.DisplayData()
            Task.sel.clear()
            Task_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Task_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Goal:
    def __init__(self, top=None):
        top.geometry("1200x700")
        top.resizable(1, 1)
        top.title("Administración - Metas")
        top.state("zoomed")

        self.message = Label(top)
        self.message.place(relx=0.035, rely=0.050, width=100, height=30)
        self.message.configure(font="-family {Poppins} -size 12")
        self.message.configure(text=username)
        self.message.configure(anchor="w")

        self.clock = Label(goals)
        self.clock.place(relx=0.9, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        self.entry1 = Entry(goals)
        self.entry1.place(relx=0.040, rely=0.286, width=200, height=35)
        self.entry1.configure(
            font="-family {Poppins} -size 12",
            relief="solid",
            bg="#f0f0f0",
            fg="#333333",
        )

        self.button1 = ttk.Button(
            top,
            text="Buscar ID",
            style="info-outline.TButton",
            command=self.search_trans,
        )
        self.button1.place(relx=0.229, rely=0.286, width=76, height=35)
        self.button1.configure(cursor="hand2")

        self.button_export_excel = ttk.Button(
            top,
            text="Exportar a excel",
            style="success.TButton",
            command=self.export_excel,
        )
        self.button_export_excel.place(relx=0.7, rely=0.106, width=250, height=35)
        self.button_export_excel.configure(cursor="hand2")

        icon_path_back = "img/flecha_atras24.png"

        self.icon_back = PhotoImage(file=icon_path_back)

        self.button2 = ttk.Button(
            top, image=self.icon_back, style="danger.TButton", command=self.Exit
        )
        self.button2.place(relx=0.035, rely=0.106, width=76, height=35)
        self.button2.configure(cursor="hand2")

        self.button3 = ttk.Button(
            top,
            text="Agregar Metas",
            style="info-outline.TButton",
            command=self.add_sent,
        )
        self.button3.place(relx=0.052, rely=0.432, width=250, height=35)
        self.button3.configure(cursor="hand2")

        self.button4 = ttk.Button(
            top,
            text="Actualizar Metas",
            style="success-outline.TButton",
            command=self.update_inves,
        )
        self.button4.place(relx=0.052, rely=0.5, width=250, height=35)
        self.button4.configure(cursor="hand2")

        self.button5 = ttk.Button(
            top,
            text="Eliminar Metas",
            style="danger-outline.TButton",
            command=self.delete_project,
        )
        self.button5.place(relx=0.052, rely=0.57, width=250, height=35)
        self.button5.configure(cursor="hand2")

        self.scrollbarx = Scrollbar(goals, orient=HORIZONTAL)
        self.scrollbary = Scrollbar(goals, orient=VERTICAL)
        self.tree = ttk.Treeview(goals)
        self.tree.place(relx=0.307, rely=0.203, width=880, height=500)
        self.tree.configure(
            yscrollcommand=self.scrollbary.set, xscrollcommand=self.scrollbarx.set
        )
        self.tree.configure(selectmode="extended")

        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.scrollbary.configure(command=self.tree.yview)
        self.scrollbarx.configure(command=self.tree.xview)

        self.scrollbary.place(relx=0.954, rely=0.203, width=22, height=500)
        self.scrollbarx.place(relx=0.307, rely=0.924, width=884, height=22)

        self.tree.configure(
            columns=(
                "ID",
                "Name",
                "Description",
                "Status",
                "Start_date",
                "End_date"
            )
        )

        self.tree.heading("ID", text="ID", anchor=W)
        self.tree.heading("Name", text="Nombre", anchor=W)
        self.tree.heading("Description", text="Descripción", anchor=W)
        self.tree.heading("Status", text="Estado", anchor=W)
        self.tree.heading("Start_date", text="Fecha de Inicio", anchor=W)
        self.tree.heading("End_date", text="Fecha de termino", anchor=W)

        self.tree.column("#0", stretch=NO, minwidth=0, width=0)
        self.tree.column("#1", stretch=NO, minwidth=0, width=80)
        self.tree.column("#2", stretch=NO, minwidth=0, width=260)
        self.tree.column("#3", stretch=NO, minwidth=0, width=100)
        self.tree.column("#4", stretch=NO, minwidth=0, width=260)
        self.tree.column("#5", stretch=NO, minwidth=0, width=100)
        self.DisplayData()

    def DisplayData(self):
        cur.execute("SELECT * FROM business_goals")
        fetch = cur.fetchall()
        for data in fetch:
            self.tree.insert("", "end", values=(data))

    def search_trans(self):
        val = []
        for i in self.tree.get_children():
            val.append(i)
            for j in self.tree.item(i)["values"]:
                val.append(j)

        try:
            to_search = int(self.entry1.get())
        except ValueError:
            messagebox.showerror("Error", "Id de la transacción invalida.", parent=taskk)
        else:
            for search in val:
                if search == to_search:
                    self.tree.selection_set(val[val.index(search) - 1])
                    self.tree.focus(val[val.index(search) - 1])
                    messagebox.showinfo(
                        "Mensaje",
                        "ID de la Tarea: {} encontrado.".format(self.entry1.get()),
                        parent=inves,
                    )
                    break
            else:
                messagebox.showerror(
                    "Error",
                    "ID Tarea: {} not found.".format(self.entry1.get()),
                    parent=inves,
                )

    sel = []

    def on_tree_select(self, Event):
        self.sel.clear()
        for i in self.tree.selection():
            if i not in self.sel:
                self.sel.append(i)

    def delete_project(self):
            selected_item = self.tree.focus()  # Obtener el item seleccionado en la tabla
            if not selected_item:
                messagebox.showwarning("Advertencia", "Seleccione una inversión para eliminar.")
                return

            # Obtener el ID y otros datos del cliente seleccionado
            client_data = self.tree.item(selected_item)['values']
            inv_id = client_data[0]  # ID del cliente

            # Mostrar confirmación antes de eliminar
            confirm = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de eliminar al cliente {client_data[1]} {client_data[2]}?"
            )
            if confirm:
                try:
                    cur.execute("DELETE FROM business_goals WHERE id = ?", (inv_id,))
                    db.commit()
                    messagebox.showinfo("Éxito", "Meta empresarial eliminada correctamente.")
                    page_goal.tree.delete(*page_goal.tree.get_children())
                    self.DisplayData()  # Actualizar la tabla después de eliminar
                except sqlite3.Error as e:
                    messagebox.showerror("Error", f"Error al eliminar la meta: {e}")

    def update_inves(self):
        if len(self.sel) == 1:
            global Goal_update
            Goal_update = Toplevel()
            page_goalup = Update_Goal(Goal_update)
            page_goalup.time()
            Goal_update.protocol("WM_DELETE_WINDOW", self.ex2)
            global valll
            valll = []
            for i in self.sel:
                for j in self.tree.item(i)["values"]:
                    valll.append(j)

            page_goalup.entry_id.insert(0, valll[0])
            page_goalup.entry1.insert(0, valll[1])
            page_goalup.entry2.insert(0, valll[2])
            page_goalup.entry3.insert(0, valll[3])
            page_goalup.entry4.insert(0, valll[4])
            page_goalup.entry5.insert(0, valll[5])
            Goal_update.mainloop()

        elif len(self.sel) == 0:
            messagebox.showerror(
                "Error", "Por favor seleccione el producto en la tabla", parent=goals
            )
        else:
            messagebox.showerror(
                "Error", "Solo puede seleccionar un producto", parent=goals
            )

    def add_sent(self):
        global page_add_goals
        global pagegoal
        page_add_goals = Toplevel()
        pagegoal = add_goal(page_add_goals)
        pagegoal.time()
        page_add_goals.mainloop()

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

    def Exit(self):
        sure = messagebox.askyesno("Panel de control", "¿Quiere regresar al panel de control?", parent=goals)
        if sure == True:
            goals.destroy()
            adm.deiconify()

    def ex2(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de salir?", parent=Goal_update)
        if sure == True:
            Goal_update.destroy()
            inv.deiconify()

    def Logout(self):
        sure = messagebox.askyesno("Salir", "¿Está seguro de cerrar la sesión?")
        if sure == True:
            root.deiconify()
            page1.entry1.delete(0, END)
            page1.entry2.delete(0, END)

    def export_excel(self):
        try:
            # Consulta SQL para obtener los datos que deseas exportar
            query = "SELECT * FROM business_goals"
            cur.execute(query)
            data = cur.fetchall()

            # Crear un nuevo libro de Excel y una hoja de cálculo
            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos SQLite'

            # Escribir los encabezados de las columnas en la hoja de cálculo
            column_names = [description[0] for description in cur.description]
            ws.append(column_names)

            # Escribir los datos de la base de datos en la hoja de cálculo
            for row in data:
                ws.append(row)

            # Guardar el archivo Excel
            wb.save('excel/metas_empresariales.xlsx')
            
            messagebox.showinfo("Éxito", "Datos exportados correctamente a Excel.")

        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Error al exportar los datos: {e}")

        finally:
            if connect_to_database():
                connect_to_database().close()

class add_goal:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Agregar Metas")

        # self.label1 = Label(p_add)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./add_product.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(page_add_goals)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            top, text="Registro de Metas Empresariales", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        # Etiqueta para rut del proveedor
        self.label_rut = Label(top, text="Nombre", font=("Times", 15), foreground="#fff")
        self.label_rut.place(relx=0.132, rely=0.2555)

        self.entry1 = ttk.Entry(page_add_goals, bootstyle="")
        self.entry1.place(relx=0.132, rely=0.296, width=730, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")

        # Etiqueta para categoría del producto
        self.label_name = Label(
            top, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_name.place(relx=0.132, rely=0.370)

        self.entry2 = ttk.Entry(page_add_goals, bootstyle="danger")
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")

        # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.132, rely=0.480)

        self.entry3 = ttk.Entry(page_add_goals, bootstyle="danger")
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")

         # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Fecha de Inicio", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.527, rely=0.255)

        
        #self.button1 = ttk.Button(
        #    top,
        #    text="Buscar ID",
        #    style="info-outline.TButton",
        #    command=self.search_trans,
        #)

        self.entry4 = ttk.Entry(page_add_goals, bootstyle="danger")
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")

         # Etiqueta para stock del producto
        self.label_email = Label(
            top, text="Fecha de Termino", font=("Times", 15), foreground="#fff"
        )

        self.label_email.place(relx=0.527, rely=0.375)

        self.entry5 = ttk.Entry(page_add_goals, bootstyle="danger")
        self.entry5.place(relx=0.527, rely=0.413, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")

        self.button1 = ttk.Button(
            top, text="Agregar", style="info.TButton", command=self.add
        )
        self.button1.place(relx=0.408, rely=0.836, width=86, height=34)
        self.button1.configure(cursor="hand2")

        self.button2 = ttk.Button(
            top, text="Limpiar", style="danger-outline.TButton", command=self.clearr
        )
        self.button2.place(relx=0.526, rely=0.836, width=86, height=34)
        self.button2.configure(cursor="hand2")

    def add(self):
        name = self.entry1.get()
        description = self.entry2.get()
        status = self.entry3.get()
        start_date = self.entry4.get()
        end_date = self.entry5.get()


        if name.strip():
            if description.strip():
                    insert = "INSERT INTO business_goals(name, description, status, start_date, end_date) VALUES(?,?,?,?,?)"
                    cur.execute(
                        insert,
                        [
                            name,
                            description,
                            status,
                            start_date,
                            end_date
                        ],
                    )
                    db.commit()
                    messagebox.showinfo(
                        "Información",
                        "Tarea ingresada correctamente",
                        parent=page_add_goals,
                    )
                    page_add_goals.destroy()
                    page_goal.tree.delete(*page_goal.tree.get_children())
                    page_goal.DisplayData()
                    page_add_goals.destroy()

            else:
                messagebox.showerror(
                    "Error", "Por favor ingrese sub categoría", parent=page_add_goals
                )
        else:
            messagebox.showerror("Error", "Por favor ingrese nombre.", parent=page_add_goals)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry6.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

class Update_Goal:
    def __init__(self, top=None):
        top.geometry("900x700")
        top.resizable(0, 0)
        top.title("Actualizar Metas")

        # self.label1 = Label(e_update)
        # self.label1.place(relx=0, rely=0, width=1366, height=768)
        # self.img = PhotoImage(file="./update_employee.png")
        # self.label1.configure(image=self.img)

        self.clock = Label(Goal_update)
        self.clock.place(relx=0.84, rely=0.065, width=102, height=36)
        self.clock.configure(font="-family {Poppins Light} -size 12")

        # Etiqueta para nombre del producto
        self.label_form = Label(
            Goal_update, text="Actualizar Tarea", font=("Times", 20), foreground="#fff"
        )
        self.label_form.place(relx=0.132, rely=0.065)

        self.entry_id = Entry(Goal_update)
        self.entry_id.place(relx=0.132, rely=0.155, width=374, height=30)
        self.entry_id.configure(font="-family {Poppins} -size 12")


        # Etiqueta para nombre del empleado
        self.label_nombre = Label(
            Goal_update, text="Nombre", font=("Times", 15), foreground="#fff"
        )
        self.label_nombre.place(relx=0.132, rely=0.2555)

        self.entry1 = Entry(Goal_update)
        self.entry1.place(relx=0.132, rely=0.296, width=374, height=30)
        self.entry1.configure(font="-family {Poppins} -size 12")
        self.entry1.configure(relief="flat")

        # Etiqueta para contacto
        self.label_contacto = Label(
            Goal_update, text="Descripción", font=("Times", 15), foreground="#fff"
        )
        self.label_contacto.place(relx=0.132, rely=0.375)

        self.entry2 = Entry(Goal_update)
        self.entry2.place(relx=0.132, rely=0.413, width=374, height=30)
        self.entry2.configure(font="-family {Poppins} -size 12")
        self.entry2.configure(relief="flat")
        #self.entry2.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Goal_update, text="Estado", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.132, rely=0.481)

        self.entry3 = Entry(Goal_update)
        self.entry3.place(relx=0.132, rely=0.529, width=374, height=30)
        self.entry3.configure(font="-family {Poppins} -size 12")
        self.entry3.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Goal_update, text="Fecha de Inicio", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.527, rely=0.255)

        self.entry4 = Entry(Goal_update)
        self.entry4.place(relx=0.527, rely=0.296, width=374, height=30)
        self.entry4.configure(font="-family {Poppins} -size 12")
        self.entry4.configure(relief="flat")
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        # Etiqueta para código
        self.label_designation = Label(
            Goal_update, text="Fecha de Termino", font=("Times", 15), foreground="#fff"
        )

        self.label_designation.place(relx=0.527, rely=0.375)


        self.entry5 = Entry(Goal_update)
        self.entry5.place(relx=0.527, rely=0.375, width=374, height=30)
        self.entry5.configure(font="-family {Poppins} -size 12")
        self.entry5.configure(relief="flat")
    
        #self.entry3.configure(validate="key", validatecommand=(self.r1, "%P"))

        self.button1 = ttk.Button(
            Goal_update, text="Actualizar", style="warning.TButton", command=self.update
        )
        self.button1.place(relx=0.408, rely=0.836, width=96, height=34)
        self.button1.configure(cursor="hand2")

    def update(self):
        id_goal = self.entry_id.get()
        name = self.entry1.get()
        description = self.entry2.get()
        status = self.entry3.get()
        start_date = self.entry4.get()
        end_date = self.entry5.get()

        if name.strip() and description and status:
            update_query = "UPDATE business_goals SET name = ?, description = ?, status = ?, start_date = ?, end_date = ? WHERE id = ?"
            cur.execute(
                update_query,
                [name, description, status, start_date, end_date, id_goal],
            )
            db.commit()
            messagebox.showinfo(
                "Información",
                "Envío ID: {} Actualizado correctamenete".format(id_goal),
                parent=Goal_update,
            )
            page_goal.tree.delete(*page_goal.tree.get_children())
            page_goal.DisplayData()
            Goal.sel.clear()
            Goal_update.destroy()
        else:
            messagebox.showerror("Oops!", "Please fill in all fields.", parent=Goal_update)

    def clearr(self):
        self.entry1.delete(0, END)
        self.entry2.delete(0, END)
        self.entry3.delete(0, END)
        self.entry4.delete(0, END)
        self.entry5.delete(0, END)

    def testint(self, val):
        if val.isdigit():
            return True
        elif val == "":
            return True
        return False

    def testchar(self, val):
        if val.isalpha():
            return True
        elif val == "":
            return True
        return False

    def time(self):
        string = strftime("%H:%M:%S %p")
        self.clock.config(text=string)
        self.clock.after(1000, self.time)

page1 = login_page(root)
root.bind("<Return>", login_page.login)
root.mainloop()
