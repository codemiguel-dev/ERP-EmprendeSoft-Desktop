import os

import pandas as pd
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controlleraddress import AddressController
from view.admin.address.viewadd import Viewadd
from view.admin.address.viewupdate import Viewupdate
from view.admin.maps.viewmaps import MapaApp


class Viewmainaddress(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmainaddress, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/admin/mainaddress{self.theme}.ui", self)

        icon_configurate_manager(self)
        icon_configurate_top(self)
        icon_exit_program(self)
        icon_excel(self)
        delete_banner(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))

        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show)
        self.btn_update.clicked.connect(self.update)
        self.btn_delete.clicked.connect(self.delete)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_maps.clicked.connect(self.maps)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = AddressController(self)

        self.show()

    def close_program(self):
        QApplication.quit()

    def reload_ui(self):
        self.close()  # Cierra la ventana actual
        self.__init__()  # Re-inicia la vista con el nuevo tema
        self.show()  # Vuelve a mostrar la ventana

    def add(self):
        self.address_add = Viewadd()
        self.address_add.show()

    def maps(self):
        """Obtiene las coordenadas de latitud y longitud para una dirección dada y muestra el mapa."""
        geolocator = Nominatim(user_agent="mi_aplicacion_geocodificador")
        address = "Avenida Ignacio Carrera Pinto 397, Illapel, Coquimbo, Chile"

        try:
            # Geocodificar la dirección
            location = geolocator.geocode(address)

            # Verificar si se obtuvo una ubicación
            if location:
                print(f"Coordenadas de '{address}':")
                print(f"Latitud: {location.latitude}, Longitud: {location.longitude}")
                self.lat = location.latitude
                self.lon = location.longitude

                # Inicializar y mostrar el mapa con las coordenadas
                self.map = MapaApp(self.lat, self.lon)
                self.map.show()
            else:
                print(f"No se encontró la dirección: {address}")

        except Exception as e:
            print(f"Ocurrió un error durante la geocodificación: {e}")

    def show(self):
        # Obtener datos desde el controlador (usando el INNER JOIN)
        Address_controller = self.controller.get()

        # Limpiar la tabla antes de insertar nuevos datos
        self.table_address.setRowCount(0)

        # Configuración del tamaño de las filas y columnas
        self.table_address.verticalHeader().setDefaultSectionSize(200)
        self.table_address.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar los datos obtenidos en la tabla
        for i, address in enumerate(Address_controller):
            self.table_address.insertRow(i)

            # Verifica si `task` es una tupla o lista con múltiples elementos
            if isinstance(address, (tuple, list)):
                # Asumiendo que la consulta INNER JOIN devuelve al menos estos campos:
                # (user_id, user_name, task_id, task_description, task_status)
                self.table_address.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(address[0]))
                )  # user_id
                self.table_address.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(address[1])
                )  # country
                self.table_address.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(address[2])
                )  # region
                self.table_address.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(address[3])
                )  # commune
                self.table_address.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(address[4])
                )  # description
            else:
                # Si `task` no es una tupla o lista, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {address}"
                )

    def update(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_address.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_address.item(selected_row, 0)
            pais_item = self.table_address.item(selected_row, 1)
            region_item = self.table_address.item(selected_row, 2)
            commune_item = self.table_address.item(selected_row, 3)
            descripcion_item = self.table_address.item(selected_row, 4)

            if (
                uid_item
                and pais_item
                and region_item
                and commune_item
                and descripcion_item
            ):
                uid = uid_item.text()
                pais = pais_item.text()
                region = region_item.text()
                commune = commune_item.text()
                descripcion = descripcion_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(uid, pais, region, commune, descripcion)
                self.update_form.show()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def delete(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_client.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_client.item(selected_row, 0)

            if uid_item:
                uid = uid_item.text()

                # Cuadro de diálogo de confirmación
                reply = QtWidgets.QMessageBox.question(
                    self,
                    "Confirmar eliminación",
                    f"¿Estás seguro de que deseas eliminar el usuario con ID {uid}?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No,
                )

                # Si el usuario confirma, se procede a eliminar
                if reply == QtWidgets.QMessageBox.Yes:
                    self.controller.delete_client(uid)
                    self.show_client()
                else:
                    # Cancelar la eliminación
                    QtWidgets.QMessageBox.information(
                        self, "Cancelado", "La eliminación ha sido cancelada."
                    )

        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        address_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            address_controller,
            columns=["ID", "País", "Region", "Comuna", "Descripción"],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_address.rowCount()):
            match = False
            for col in range(self.table_address.columnCount()):
                item = self.table_address.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_address.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
