import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QSizeGrip
from PyQt5.uic import loadUi

from configuration.configuration_buttom import (
    icon_configurate_manager,
    icon_configurate_top,
    icon_excel,
    icon_exit_program,
)
from configuration.configuration_buttom_top import (
    control_bt_maximizar,
    control_bt_minimizar,
    control_bt_normal,
)
from configuration.configuration_config_theme import load_config
from configuration.configuration_delete_banner import delete_banner
from configuration.configuration_window_move import mousePressEvent, window_move
from controller.controllertask import TaskController
from view.admin.task.viewadd import Viewadd
from view.admin.task.viewupdate import Viewupdate


class Viewmaintask(QtWidgets.QMainWindow):
    def __init__(self):
        super(Viewmaintask, self).__init__()
        self.theme = load_config(self)  # Lee la configuración al iniciar
        loadUi(f"design/admin/maintask{self.theme}.ui", self)

        delete_banner(self)
        icon_configurate_manager(self)
        icon_configurate_top(self)
        icon_exit_program(self)
        icon_excel(self)

        # Inicializar el QSizeGrip y establecer su tamaño
        self.gripSize = 16  # Define el tamaño del grip
        self.grip = QSizeGrip(self)  # Crea un QSizeGrip para el redimensionamiento
        self.grip.resize(self.gripSize, self.gripSize)  # Ajusta el tamaño del QSizeGrip

        # Maximizar la ventana por defecto
        self.showMaximized()

        # Asigna los eventos personalizados al frame superior
        self.frame_superior.mousePressEvent = lambda event: mousePressEvent(self, event)
        self.frame_superior.mouseMoveEvent = lambda event: window_move(self, event)

        # Control de botones de la barra de títulos
        self.bt_minimizar.clicked.connect(lambda: control_bt_minimizar(self))
        self.bt_restaurar.clicked.connect(lambda: control_bt_normal(self))
        self.bt_maximizar.clicked.connect(lambda: control_bt_maximizar(self))
        self.bt_cerrar.clicked.connect(lambda: self.close())
        self.bt_maximizar.hide()

        self.btn_add.clicked.connect(self.add)
        self.btn_get.clicked.connect(self.show)
        self.btn_update.clicked.connect(self.update)
        self.btn_delete.clicked.connect(self.delete)
        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_search.clicked.connect(self.search)
        self.btn_exit.clicked.connect(self.close_program)

        self.controller = TaskController(self)

        self.show()

    def close_program(self):
        QApplication.quit()

    def add(self):
        self.task_add = Viewadd()
        self.task_add.show()

    def show(self):
        # Obtener datos desde el controlador (usando el INNER JOIN)
        Task_controller = self.controller.get()

        # Limpiar la tabla antes de insertar nuevos datos
        self.table_task.setRowCount(0)

        # Configuración del tamaño de las filas y columnas
        self.table_task.verticalHeader().setDefaultSectionSize(200)
        self.table_task.horizontalHeader().setDefaultSectionSize(300)

        # Mostrar los datos obtenidos en la tabla
        for i, task in enumerate(Task_controller):
            self.table_task.insertRow(i)

            # Verifica si `task` es una tupla o lista con múltiples elementos
            if isinstance(task, (tuple, list)):
                # Asumiendo que la consulta INNER JOIN devuelve al menos estos campos:
                # (user_id, user_name, task_id, task_description, task_status)
                self.table_task.setItem(
                    i, 0, QtWidgets.QTableWidgetItem(str(task[0]))
                )  # user_id
                self.table_task.setItem(
                    i, 1, QtWidgets.QTableWidgetItem(task[1])
                )  # user_name
                self.table_task.setItem(
                    i, 2, QtWidgets.QTableWidgetItem(str(task[2]))
                )  # task_id
                self.table_task.setItem(
                    i, 3, QtWidgets.QTableWidgetItem(task[3])
                )  # task_name
                self.table_task.setItem(
                    i, 4, QtWidgets.QTableWidgetItem(task[4])
                )  # task_description
                self.table_task.setItem(
                    i, 5, QtWidgets.QTableWidgetItem(task[5])
                )  # task_status
            else:
                # Si `task` no es una tupla o lista, maneja el error
                print(
                    f"Error: Se esperaba una tupla, pero se encontró un valor único: {task}"
                )

    def update(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_task.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_task.item(selected_row, 0)
            user_item = self.table_task.item(selected_row, 1)
            id_task_item = self.table_task.item(selected_row, 2)
            name_item = self.table_task.item(selected_row, 3)
            description_item = self.table_task.item(selected_row, 4)
            status_item = self.table_task.item(selected_row, 5)

            if (
                uid_item
                and user_item
                and id_task_item
                and name_item
                and description_item
                and status_item
            ):
                uid = uid_item.text()
                user = user_item.text()
                id_task = id_task_item.text()
                name = name_item.text()
                description = description_item.text()
                status = status_item.text()

                # Abrir el nuevo formulario de actualización
                self.update_form = Viewupdate(
                    uid, id_task, user, name, description, status
                )
                self.update_form.show()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def delete(self):
        # Obtener índice de la fila seleccionada
        selected_row = self.table_task.currentRow()

        # Verificar si se ha seleccionado una fila
        if selected_row != -1:
            uid_item = self.table_task.item(selected_row, 2)

            if uid_item:
                uid = uid_item.text()

                # Cuadro de diálogo de confirmación
                reply = QtWidgets.QMessageBox.question(
                    self,
                    "Confirmar eliminación",
                    f"¿Estás seguro de que deseas eliminar el usuario con ID {uid}?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.No,
                )

                # Si el usuario confirma, se procede a eliminar
                if reply == QtWidgets.QMessageBox.Yes:
                    self.controller.delete(uid)
                    self.show()
                else:
                    # Cancelar la eliminación
                    QtWidgets.QMessageBox.information(
                        self, "Cancelado", "La eliminación ha sido cancelada."
                    )

        else:
            QtWidgets.QMessageBox.warning(
                self, "Error", "Por favor, seleccione una fila para actualizar."
            )

    def export_excel(self):
        # Obtener datos de usuarios desde el controlador
        task_controller = self.controller.get()

        # Convertir los datos a un DataFrame de pandas
        df = pd.DataFrame(
            task_controller,
            columns=[
                "ID del usuario",
                "Nombre del usuario",
                "ID de la tarea",
                "Nombre de la tarea",
                "Descripción de la tarea",
                "Estado de la tarea",
            ],
        )

        # Mostrar el cuadro de diálogo para guardar el archivo
        options = QFileDialog.Options()
        initial_directory = os.path.join(
            os.path.dirname(__file__), "../../../excel"
        )  # Directorio inicial
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar",
            initial_directory,
            "Archivos de Excel (*.xlsx);;Todos los archivos (*)",
            options=options,
        )

        if file_path:
            # Guardar el DataFrame en un archivo de Excel
            df.to_excel(file_path, index=False)

            # Cargar el archivo Excel para aplicar estilos
            wb = load_workbook(file_path)
            ws = wb.active

            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primera fila contiene los encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Guardar los cambios en el archivo
            wb.save(file_path)

            QtWidgets.QMessageBox.information(
                self, "Éxito", f"Usuario exportado correctamente a {file_path}"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, "Cancelado", "La exportación fue cancelada."
            )

    def search(self):
        search_text = self.searchtxt.text().lower()  # Texto de búsqueda en minúsculas
        found = False  # Variable para rastrear si se encontró una coincidencia

        for row in range(self.table_task.rowCount()):
            match = False
            for col in range(self.table_task.columnCount()):
                item = self.table_task.item(row, col)
                if item and search_text in item.text().lower():  # Comparar texto
                    match = True
                    found = True  # Si hay una coincidencia, actualizamos 'found'
                    break  # Si encuentra coincidencia, no sigue buscando en esa fila

            # Ocultar la fila si no coincide
            self.table_task.setRowHidden(row, not match)

        # Si no se encontró ninguna coincidencia, mostrar un mensaje de advertencia
        if not found:
            QtWidgets.QMessageBox.warning(
                self,
                "Sin coincidencias",
                "No se encontraron resultados para la búsqueda.",
            )
